"""
השאלה חכמה - מערכת ניהול גמ"ח חכמה
Smart Lending - Gemach Management System
Version 1.5
"""

import sys
import os
import sqlite3
import smtplib
import threading
from datetime import datetime, timedelta, date
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
import base64
import json

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QPushButton, QTableWidget, QTableWidgetItem, QDialog,
    QLineEdit, QComboBox, QDateEdit, QTextEdit, QFormLayout,
    QStackedWidget, QFrame, QScrollArea, QMessageBox, QFileDialog,
    QTabWidget, QCalendarWidget, QSplitter, QHeaderView, QAbstractItemView,
    QProgressBar, QToolButton, QMenu, QSystemTrayIcon, QGroupBox,
    QCheckBox, QSpinBox, QStatusBar, QSizePolicy, QCompleter
)
from PySide6.QtCore import (
    Qt, QDate, QTimer, QThread, Signal, QSortFilterProxyModel,
    QStringListModel, QSize, QPropertyAnimation, QEasingCurve, QPoint, QRect,
    QObject, Slot, QUrl
)
from PySide6.QtGui import (
    QFont, QColor, QPalette, QIcon, QPixmap, QPainter, QBrush,
    QLinearGradient, QFontDatabase, QAction, QCursor
)

# Optional rich-dashboard support (HTML/CSS). Falls back to the classic
# Qt-widgets dashboard automatically if QtWebEngine isn't available.
try:
    from PySide6.QtWebEngineWidgets import QWebEngineView
    from PySide6.QtWebChannel import QWebChannel
    WEBENGINE_AVAILABLE = True
except Exception:
    WEBENGINE_AVAILABLE = False


# ─────────────────────────────────────────────
#  CONSTANTS & THEME
# ─────────────────────────────────────────────

DB_PATH = Path.home() / "hashala_chachama.db"
IMAGES_DIR = Path.home() / "hashala_images"
IMAGES_DIR.mkdir(exist_ok=True)

THEMES = {
    "navy_gold": {
        "display_name": "כחול נייבי וזהב (מקורי)",
        "primary":      "#1B4F72",
        "primary_light":"#2E86C1",
        "accent":       "#27AE60",
        "warning":      "#F39C12",
        "danger":       "#E74C3C",
        "surface":      "#FFFFFF",
        "background":   "#F0F4F8",
        "sidebar":      "#1B2A3B",
        "sidebar_hover":"#243447",
        "sidebar_active":"#2E86C1",
        "text_primary": "#1A202C",
        "text_secondary":"#718096",
        "border":       "#E2E8F0",
        "card":         "#FFFFFF",
        "badge_blue":   "#EBF8FF",
        "badge_green":  "#F0FFF4",
        "badge_red":    "#FFF5F5",
        "badge_yellow": "#FFFFF0",
        "html_navy": "#0F1B2E", "html_navy2": "#16263F", "html_gold": "#D4A24C",
        "html_gold_light": "#E8C588", "html_sage": "#5B8C6E", "html_sage_light": "#E3EEE6",
        "html_rose": "#C75D4D", "html_rose_light": "#F7E2DE", "html_parchment": "#F7F3EC",
        "html_line": "#E7E0D2",
    },
    "forest": {
        "display_name": "ירוק יערני וחום אדמה",
        "primary":      "#1F4E3D",
        "primary_light":"#2F7A5C",
        "accent":       "#8A9A3B",
        "warning":      "#C68642",
        "danger":       "#B14B3D",
        "surface":      "#FFFFFF",
        "background":   "#F2F4EE",
        "sidebar":      "#16302A",
        "sidebar_hover":"#1F3F37",
        "sidebar_active":"#2F7A5C",
        "text_primary": "#22281F",
        "text_secondary":"#6E776A",
        "border":       "#DCE3D6",
        "card":         "#FFFFFF",
        "badge_blue":   "#EAF3EE",
        "badge_green":  "#F1F6E6",
        "badge_red":    "#FBEEE9",
        "badge_yellow": "#FBF2E6",
        "html_navy": "#16302A", "html_navy2": "#1F3F37", "html_gold": "#C68642",
        "html_gold_light": "#E0B07F", "html_sage": "#8A9A3B", "html_sage_light": "#EEF1DD",
        "html_rose": "#B14B3D", "html_rose_light": "#F3DFD8", "html_parchment": "#F2F4EE",
        "html_line": "#DCE3D6",
    },
    "slate_rose": {
        "display_name": "אפור פחם וורוד עתיק",
        "primary":      "#3B3C53",
        "primary_light":"#6C6E94",
        "accent":       "#7A8B69",
        "warning":      "#C9974C",
        "danger":       "#B5566B",
        "surface":      "#FFFFFF",
        "background":   "#F4F2F3",
        "sidebar":      "#26273A",
        "sidebar_hover":"#34354C",
        "sidebar_active":"#6C6E94",
        "text_primary": "#272733",
        "text_secondary":"#7A7A8C",
        "border":       "#E5E1E6",
        "card":         "#FFFFFF",
        "badge_blue":   "#EFEFF6",
        "badge_green":  "#EEF2EA",
        "badge_red":    "#F8EBEE",
        "badge_yellow": "#FAF1E3",
        "html_navy": "#26273A", "html_navy2": "#34354C", "html_gold": "#C9974C",
        "html_gold_light": "#DFB87E", "html_sage": "#7A8B69", "html_sage_light": "#EAEFE4",
        "html_rose": "#B5566B", "html_rose_light": "#F3DEE3", "html_parchment": "#F4F2F3",
        "html_line": "#E5E1E6",
    },
    "midnight_dark": {
        "display_name": "כהה (Dark Mode)",
        "primary":      "#3B82C4",
        "primary_light":"#5BA3E0",
        "accent":       "#4CAF7D",
        "warning":      "#E0A858",
        "danger":       "#E0716A",
        "surface":      "#1E2433",
        "background":   "#15192480",
        "background_solid": "#151924",
        "sidebar":      "#0E1119",
        "sidebar_hover":"#1A1F2C",
        "sidebar_active":"#3B82C4",
        "text_primary": "#E8EAF0",
        "text_secondary":"#8D94A8",
        "border":       "#2B3142",
        "card":         "#1E2433",
        "badge_blue":   "#1C2C3D",
        "badge_green":  "#1B3329",
        "badge_red":    "#3A2228",
        "badge_yellow": "#3A301E",
        "html_navy": "#0E1119", "html_navy2": "#1A1F2C", "html_gold": "#E0A858",
        "html_gold_light": "#EAC287", "html_sage": "#4CAF7D", "html_sage_light": "#1B3329",
        "html_rose": "#E0716A", "html_rose_light": "#3A2228", "html_parchment": "#151924",
        "html_line": "#2B3142",
    },
}

DEFAULT_THEME = "navy_gold"
COLORS = dict(THEMES[DEFAULT_THEME])


def build_stylesheet(colors):
    """Build the global Qt stylesheet from a theme color dict."""
    bg = colors.get('background_solid', colors['background'])
    return f"""
QMainWindow, QWidget#centralWidget {{
    background-color: {bg};
    font-family: 'Segoe UI', Arial;
}}

/* ── Global RTL text alignment ── */
QLabel {{
    qproperty-alignment: 'AlignRight|AlignVCenter';
    color: {colors['text_primary']};
}}

/* ── Sidebar ── */
QWidget#sidebar {{
    background-color: {colors['sidebar']};
    border-left: none;
}}
QFrame#logoFrame {{
    background-color: {colors['primary']};
    padding: 0;
}}
QPushButton#navBtn {{
    background: transparent;
    color: #A0AEC0;
    border: none;
    padding: 14px 20px 14px 16px;
    text-align: right;
    font-size: 14px;
    font-weight: 500;
    border-radius: 0;
    border-left: 4px solid transparent;
}}
QPushButton#navBtn:hover {{
    background-color: {colors['sidebar_hover']};
    color: #FFFFFF;
}}
QPushButton#navBtn[active=true] {{
    background-color: {colors['sidebar_active']};
    color: #FFFFFF;
    border-left: 4px solid #74B9FF;
}}

/* ── Cards ── */
QFrame#card {{
    background: {colors['card']};
    border-radius: 12px;
    border: 1px solid {colors['border']};
}}

/* ── Stat Cards ── */
QFrame#statCard {{
    background: {colors['card']};
    border-radius: 12px;
    border: 1px solid {colors['border']};
    padding: 4px;
}}

/* ── Tables ── */
QTableWidget {{
    background: {colors['card']};
    border: 1px solid {colors['border']};
    border-radius: 8px;
    gridline-color: {colors['border']};
    font-size: 13px;
    color: {colors['text_primary']};
    outline: none;
}}
QTableWidget::item {{
    padding: 10px 14px;
    border-bottom: 1px solid {colors['border']};
}}
QTableWidget::item:selected {{
    background-color: {colors['badge_blue']};
    color: {colors['primary']};
}}
QHeaderView::section {{
    background-color: {colors['card']};
    color: {colors['text_secondary']};
    padding: 10px 14px;
    font-size: 12px;
    font-weight: 600;
    border: none;
    border-bottom: 2px solid {colors['border']};
    text-align: right;
}}
QHeaderView {{
    background-color: {colors['card']};
}}

/* ── Inputs ── */
QLineEdit, QTextEdit, QComboBox, QDateEdit, QSpinBox {{
    background: {colors['surface']};
    border: 1.5px solid {colors['border']};
    border-radius: 8px;
    padding: 9px 14px;
    font-size: 13px;
    color: {colors['text_primary']};
    min-height: 20px;
}}
QLineEdit:focus, QTextEdit:focus, QComboBox:focus, QDateEdit:focus {{
    border-color: {colors['primary_light']};
    outline: none;
}}
QComboBox::drop-down {{
    border: none;
    padding-left: 8px;
}}
QComboBox QAbstractItemView {{
    background: {colors['surface']};
    border: 1px solid {colors['border']};
    selection-background-color: {colors['badge_blue']};
    selection-color: {colors['primary']};
    border-radius: 8px;
}}

/* ── Buttons ── */
QPushButton#primaryBtn {{
    background-color: {colors['primary_light']};
    color: white;
    border: none;
    border-radius: 8px;
    padding: 10px 22px;
    font-size: 13px;
    font-weight: 600;
}}
QPushButton#primaryBtn:hover {{
    background-color: {colors['primary']};
}}
QPushButton#primaryBtn:pressed {{
    background-color: {colors['primary']};
}}
QPushButton#successBtn {{
    background-color: {colors['accent']};
    color: white;
    border: none;
    border-radius: 8px;
    padding: 10px 22px;
    font-size: 13px;
    font-weight: 600;
}}
QPushButton#successBtn:hover {{
    background-color: {colors['accent']};
}}
QPushButton#dangerBtn {{
    background-color: {colors['danger']};
    color: white;
    border: none;
    border-radius: 8px;
    padding: 10px 22px;
    font-size: 13px;
    font-weight: 600;
}}
QPushButton#dangerBtn:hover {{
    background-color: {colors['danger']};
}}
QPushButton#warningBtn {{
    background-color: {colors['warning']};
    color: white;
    border: none;
    border-radius: 8px;
    padding: 10px 22px;
    font-size: 13px;
    font-weight: 600;
}}
QPushButton#secondaryBtn {{
    background-color: transparent;
    color: {colors['primary_light']};
    border: 1.5px solid {colors['primary_light']};
    border-radius: 8px;
    padding: 9px 20px;
    font-size: 13px;
    font-weight: 600;
}}
QPushButton#secondaryBtn:hover {{
    background-color: {colors['badge_blue']};
}}
QPushButton#iconBtn {{
    background: transparent;
    border: none;
    padding: 6px;
    border-radius: 6px;
    font-size: 16px;
}}
QPushButton#iconBtn:hover {{
    background-color: {bg};
}}

/* ── Labels ── */
QLabel#pageTitle {{
    font-size: 22px;
    font-weight: 700;
    color: {colors['text_primary']};
}}
QLabel#sectionTitle {{
    font-size: 15px;
    font-weight: 600;
    color: {colors['text_primary']};
}}
QLabel#statValue {{
    font-size: 28px;
    font-weight: 700;
    color: {colors['text_primary']};
}}
QLabel#statLabel {{
    font-size: 12px;
    color: {colors['text_secondary']};
    font-weight: 500;
}}

/* ── Tabs ── */
QTabWidget::pane {{
    border: 1px solid {colors['border']};
    border-radius: 8px;
    background: {colors['card']};
}}
QTabBar::tab {{
    background: transparent;
    padding: 10px 20px;
    font-size: 13px;
    color: {colors['text_secondary']};
    border-bottom: 2px solid transparent;
}}
QTabBar::tab:selected {{
    color: {colors['primary_light']};
    border-bottom: 2px solid {colors['primary_light']};
    font-weight: 600;
}}
QTabBar::tab:hover {{
    color: {colors['primary']};
}}

/* ── Scrollbar ── */
QScrollBar:vertical {{
    background: transparent;
    width: 6px;
    margin: 0;
}}
QScrollBar::handle:vertical {{
    background: #CBD5E0;
    border-radius: 3px;
    min-height: 30px;
}}
QScrollBar::handle:vertical:hover {{
    background: #A0AEC0;
}}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
    height: 0;
}}
QScrollBar:horizontal {{
    height: 6px;
    background: transparent;
}}
QScrollBar::handle:horizontal {{
    background: #CBD5E0;
    border-radius: 3px;
}}

/* ── Search ── */
QLineEdit#searchBox {{
    background: {colors['surface']};
    border: 1.5px solid {colors['border']};
    border-radius: 20px;
    padding: 8px 16px 8px 36px;
    font-size: 13px;
}}

/* ── Dialog ── */
QDialog {{
    background: {bg};
}}

/* ── Status badges ── */
QLabel#badgeAvail {{
    background: {colors['badge_green']};
    color: {colors['accent']};
    border-radius: 10px;
    padding: 3px 10px;
    font-size: 11px;
    font-weight: 600;
}}
QLabel#badgeLoan {{
    background: {colors['badge_blue']};
    color: {colors['primary_light']};
    border-radius: 10px;
    padding: 3px 10px;
    font-size: 11px;
    font-weight: 600;
}}
QLabel#badgeLate {{
    background: {colors['badge_red']};
    color: {colors['danger']};
    border-radius: 10px;
    padding: 3px 10px;
    font-size: 11px;
    font-weight: 600;
}}
QLabel#badgeRepair {{
    background: {colors['badge_yellow']};
    color: {colors['warning']};
    border-radius: 10px;
    padding: 3px 10px;
    font-size: 11px;
    font-weight: 600;
}}

/* ── Progress ── */
QProgressBar {{
    background: {colors['border']};
    border-radius: 4px;
    border: none;
    height: 8px;
    text-align: center;
}}
QProgressBar::chunk {{
    background: {colors['primary_light']};
    border-radius: 4px;
}}

/* ── Calendar ── */
QCalendarWidget QToolButton {{
    color: {colors['text_primary']};
    background: transparent;
    font-size: 13px;
}}
QCalendarWidget QMenu {{
    background: {colors['surface']};
}}
QCalendarWidget QAbstractItemView:enabled {{
    color: {colors['text_primary']};
    background: {colors['surface']};
    selection-background-color: {colors['primary_light']};
    selection-color: white;
}}
"""


STYLESHEET = build_stylesheet(COLORS)



# ─────────────────────────────────────────────
#  DATABASE
# ─────────────────────────────────────────────

class Database:
    def __init__(self):
        self.conn = sqlite3.connect(str(DB_PATH))
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys = ON")
        self.create_tables()
        self.seed_demo_data()

    def create_tables(self):
        cur = self.conn.cursor()
        cur.executescript("""
        CREATE TABLE IF NOT EXISTS items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            category TEXT NOT NULL DEFAULT 'שונות',
            status TEXT NOT NULL DEFAULT 'זמין',
            serial_number TEXT,
            location TEXT,
            notes TEXT,
            image_path TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS borrowers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT NOT NULL,
            phone TEXT NOT NULL,
            email TEXT,
            address TEXT,
            rating INTEGER DEFAULT 5,
            notes TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS loans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER NOT NULL REFERENCES items(id),
            borrower_id INTEGER NOT NULL REFERENCES borrowers(id),
            loan_date DATETIME DEFAULT CURRENT_TIMESTAMP,
            planned_return DATE NOT NULL,
            actual_return DATETIME,
            status TEXT DEFAULT 'פעיל',
            notes TEXT
        );

        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        );

        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action TEXT,
            details TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        """)
        self.conn.commit()

    def seed_demo_data(self):
        cur = self.conn.cursor()
        count = cur.execute("SELECT COUNT(*) FROM items").fetchone()[0]
        if count > 0:
            return
        # Sample items
        items = [
            ("מברג חשמלי Bosch", "כלי עבודה", "זמין", "BOS-2023", "מדף א׳", "מצב טוב"),
            ("שולחן מתקפל גדול", "ציוד אירועים", "זמין", None, "מחסן", "6 כסאות נלווים"),
            ("כסאות ערימה (10 יחידות)", "ציוד אירועים", "מושאל", None, "מחסן", ""),
            ("מדחום דיגיטלי", "רפואי", "זמין", "MED-001", "ארון ב׳", ""),
            ("מסחטת פירות", "מכשירי חשמל", "בתיקון", None, "חדר כלים", "בדיקת מנוע"),
            ("מקדחה ידנית", "כלי עבודה", "זמין", None, "מדף א׳", ""),
            ("מזרן אורח", "שונות", "זמין", None, "ארון ג׳", ""),
            ("מכונת תפירה Singer", "מכשירי חשמל", "מושאל", "SNG-77", "מדף ב׳", ""),
            ("עגלת קניות", "שונות", "זמין", None, "כניסה", ""),
            ("ניקוי שטיחים - מכונה", "מכשירי חשמל", "זמין", "CLN-03", "מחסן", "כולל אביזרים"),
        ]
        cur.executemany(
            "INSERT INTO items (name,category,status,serial_number,location,notes) VALUES (?,?,?,?,?,?)",
            items
        )
        borrowers = [
            ("ישראל ישראלי", "050-1234567", "israel@gmail.com", "הרצל 1, תל אביב", 5),
            ("שרה כהן", "052-9876543", "sara@gmail.com", "ויצמן 5, רמת גן", 4),
            ("משה לוי", "054-1111222", "", "בן גוריון 12, חיפה", 3),
            ("רחל מזרחי", "058-3334455", "rachel@gmail.com", "רוטשילד 8, ת״א", 5),
        ]
        cur.executemany(
            "INSERT INTO borrowers (full_name,phone,email,address,rating) VALUES (?,?,?,?,?)",
            borrowers
        )
        # Some loans - use ISO strings to avoid Python 3.12 date adapter deprecation
        today = date.today()
        cur.execute(
            "INSERT INTO loans (item_id,borrower_id,loan_date,planned_return,status) VALUES (3,1,?,?,?)",
            ((today - timedelta(days=5)).isoformat(), (today + timedelta(days=2)).isoformat(), "פעיל")
        )
        cur.execute(
            "INSERT INTO loans (item_id,borrower_id,loan_date,planned_return,status) VALUES (8,2,?,?,?)",
            ((today - timedelta(days=10)).isoformat(), (today - timedelta(days=3)).isoformat(), "באיחור")
        )
        cur.execute("UPDATE items SET status='מושאל' WHERE id IN (3,8)")
        self.conn.commit()

    def log(self, action, details=""):
        self.conn.execute("INSERT INTO audit_log (action,details) VALUES (?,?)", (action, details))
        self.conn.commit()

    # ── Items ──
    def get_items(self, search="", category="", status=""):
        q = "SELECT * FROM items WHERE 1=1"
        p = []
        if search:
            q += " AND name LIKE ?"
            p.append(f"%{search}%")
        if category:
            q += " AND category=?"
            p.append(category)
        if status:
            q += " AND status=?"
            p.append(status)
        q += " ORDER BY name"
        return self.conn.execute(q, p).fetchall()

    def add_item(self, name, category, status, serial, location, notes, image_path=""):
        cur = self.conn.execute(
            "INSERT INTO items (name,category,status,serial_number,location,notes,image_path) VALUES (?,?,?,?,?,?,?)",
            (name, category, status, serial, location, notes, image_path)
        )
        self.conn.commit()
        self.log("הוספת חפץ", f"{name}")
        return cur.lastrowid

    def update_item(self, item_id, name, category, status, serial, location, notes, image_path=""):
        self.conn.execute(
            "UPDATE items SET name=?,category=?,status=?,serial_number=?,location=?,notes=?,image_path=? WHERE id=?",
            (name, category, status, serial, location, notes, image_path, item_id)
        )
        self.conn.commit()
        self.log("עדכון חפץ", f"#{item_id} {name}")

    def delete_item(self, item_id):
        active = self.conn.execute(
            "SELECT COUNT(*) FROM loans WHERE item_id=? AND status='פעיל'", (item_id,)
        ).fetchone()[0]
        if active:
            return False
        self.conn.execute("DELETE FROM items WHERE id=?", (item_id,))
        self.conn.commit()
        self.log("מחיקת חפץ", f"#{item_id}")
        return True

    def get_item(self, item_id):
        return self.conn.execute("SELECT * FROM items WHERE id=?", (item_id,)).fetchone()

    # ── Borrowers ──
    def get_borrowers(self, search=""):
        q = "SELECT * FROM borrowers WHERE 1=1"
        p = []
        if search:
            q += " AND (full_name LIKE ? OR phone LIKE ?)"
            p += [f"%{search}%", f"%{search}%"]
        q += " ORDER BY full_name"
        return self.conn.execute(q, p).fetchall()

    def add_borrower(self, full_name, phone, email, address, rating, notes):
        cur = self.conn.execute(
            "INSERT INTO borrowers (full_name,phone,email,address,rating,notes) VALUES (?,?,?,?,?,?)",
            (full_name, phone, email, address, rating, notes)
        )
        self.conn.commit()
        self.log("הוספת שואל", full_name)
        return cur.lastrowid

    def update_borrower(self, bid, full_name, phone, email, address, rating, notes):
        self.conn.execute(
            "UPDATE borrowers SET full_name=?,phone=?,email=?,address=?,rating=?,notes=? WHERE id=?",
            (full_name, phone, email, address, rating, notes, bid)
        )
        self.conn.commit()
        self.log("עדכון שואל", f"#{bid} {full_name}")

    def delete_borrower(self, bid):
        active = self.conn.execute(
            "SELECT COUNT(*) FROM loans WHERE borrower_id=? AND status='פעיל'", (bid,)
        ).fetchone()[0]
        if active:
            return False
        self.conn.execute("DELETE FROM borrowers WHERE id=?", (bid,))
        self.conn.commit()
        return True

    def get_borrower(self, bid):
        return self.conn.execute("SELECT * FROM borrowers WHERE id=?", (bid,)).fetchone()

    # ── Loans ──
    def get_loans(self, active_only=False, overdue_only=False):
        q = """
        SELECT l.*, i.name as item_name, i.category,
               b.full_name as borrower_name, b.phone,
               b.email as borrower_email
        FROM loans l
        JOIN items i ON l.item_id=i.id
        JOIN borrowers b ON l.borrower_id=b.id
        WHERE 1=1
        """
        p = []
        if active_only:
            q += " AND l.status IN ('פעיל','באיחור')"
        if overdue_only:
            q += " AND l.status='באיחור'"
        q += " ORDER BY l.loan_date DESC"
        return self.conn.execute(q, p).fetchall()

    def create_loan(self, item_id, borrower_id, planned_return, notes=""):
        cur = self.conn.execute(
            "INSERT INTO loans (item_id,borrower_id,planned_return,notes) VALUES (?,?,?,?)",
            (item_id, borrower_id, planned_return, notes)
        )
        self.conn.execute("UPDATE items SET status='מושאל' WHERE id=?", (item_id,))
        self.conn.commit()
        self.log("השאלה חדשה", f"חפץ #{item_id} ← שואל #{borrower_id}")
        return cur.lastrowid

    def return_item(self, loan_id, item_id):
        now = datetime.now().isoformat()
        self.conn.execute(
            "UPDATE loans SET actual_return=?,status='הוחזר' WHERE id=?",
            (now, loan_id)
        )
        self.conn.execute("UPDATE items SET status='זמין' WHERE id=?", (item_id,))
        self.conn.commit()
        self.log("החזרת חפץ", f"השאלה #{loan_id}")

    def update_overdue_statuses(self):
        today = date.today().isoformat()
        self.conn.execute(
            "UPDATE loans SET status='באיחור' WHERE status='פעיל' AND planned_return < ?",
            (today,)
        )
        self.conn.commit()

    # ── Dashboard stats ──
    def get_stats(self):
        r = {}
        r['total'] = self.conn.execute("SELECT COUNT(*) FROM items").fetchone()[0]
        r['available'] = self.conn.execute("SELECT COUNT(*) FROM items WHERE status='זמין'").fetchone()[0]
        r['loaned'] = self.conn.execute("SELECT COUNT(*) FROM items WHERE status='מושאל'").fetchone()[0]
        r['repair'] = self.conn.execute("SELECT COUNT(*) FROM items WHERE status='בתיקון'").fetchone()[0]
        r['overdue'] = self.conn.execute("SELECT COUNT(*) FROM loans WHERE status='באיחור'").fetchone()[0]
        r['borrowers'] = self.conn.execute("SELECT COUNT(*) FROM borrowers").fetchone()[0]
        r['total_loans'] = self.conn.execute("SELECT COUNT(*) FROM loans").fetchone()[0]
        return r

    def get_overdue_loans(self):
        return self.conn.execute("""
            SELECT l.*, i.name as item_name, b.full_name as borrower_name,
                   b.phone, b.email as borrower_email
            FROM loans l
            JOIN items i ON l.item_id=i.id
            JOIN borrowers b ON l.borrower_id=b.id
            WHERE l.status='באיחור'
        """).fetchall()

    def get_audit_log(self, limit=100):
        return self.conn.execute(
            "SELECT * FROM audit_log ORDER BY timestamp DESC LIMIT ?", (limit,)
        ).fetchall()

    def get_settings(self):
        rows = self.conn.execute("SELECT key,value FROM settings").fetchall()
        s = {}
        for r in rows:
            s[r['key']] = r['value']
        return s

    def save_setting(self, key, value):
        self.conn.execute(
            "INSERT OR REPLACE INTO settings (key,value) VALUES (?,?)", (key, value)
        )
        self.conn.commit()


# ─────────────────────────────────────────────
#  HELPER WIDGETS
# ─────────────────────────────────────────────

def make_badge(text, style="avail"):
    lbl = QLabel(text)
    name_map = {"avail": "badgeAvail", "loan": "badgeLoan", "late": "badgeLate", "repair": "badgeRepair"}
    lbl.setObjectName(name_map.get(style, "badgeAvail"))
    lbl.setAlignment(Qt.AlignCenter)
    lbl.setFixedHeight(24)
    return lbl

def status_badge_style(status):
    m = {"זמין": "avail", "מושאל": "loan", "באיחור": "late", "בתיקון": "repair", "אבוד": "late"}
    return m.get(status, "avail")

def sep_line():
    f = QFrame()
    f.setFrameShape(QFrame.HLine)
    f.setStyleSheet(f"color: {COLORS['border']};")
    return f

def rtl_table(tbl):
    """Apply full RTL settings to a QTableWidget — layout direction, header
    direction, and a default right-alignment for any item that doesn't set
    its own alignment explicitly."""
    tbl.setLayoutDirection(Qt.RightToLeft)
    tbl.horizontalHeader().setLayoutDirection(Qt.RightToLeft)
    tbl.horizontalHeader().setDefaultAlignment(Qt.AlignRight | Qt.AlignVCenter)
    tbl.verticalHeader().setLayoutDirection(Qt.RightToLeft)
    return tbl

def rtl_item(text, color=None):
    """Build a right-aligned, vertically-centered QTableWidgetItem — the
    single source of truth for table-cell alignment across the app."""
    cell = QTableWidgetItem(str(text) if text is not None else "")
    cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
    if color:
        cell.setForeground(QColor(color))
    return cell

def card_frame():
    f = QFrame()
    f.setObjectName("card")
    return f

def stat_card(label, value, color=None, icon=""):
    frame = QFrame()
    frame.setObjectName("statCard")
    frame.setMinimumHeight(110)
    layout = QVBoxLayout(frame)
    layout.setContentsMargins(18, 14, 18, 14)

    top = QHBoxLayout()
    ico = QLabel(icon)
    ico.setStyleSheet(f"font-size: 24px;")
    top.addWidget(ico)
    top.addStretch()
    layout.addLayout(top)

    val = QLabel(str(value))
    val.setObjectName("statValue")
    if color:
        val.setStyleSheet(f"font-size: 28px; font-weight: 700; color: {color};")
    layout.addWidget(val)

    lbl = QLabel(label)
    lbl.setObjectName("statLabel")
    layout.addWidget(lbl)

    return frame, val


# ─────────────────────────────────────────────
#  DIALOGS
# ─────────────────────────────────────────────

CATEGORIES = ["כלי עבודה", "ציוד אירועים", "מכשירי חשמל", "רפואי", "שונות"]
ITEM_STATUSES = ["זמין", "מושאל", "בתיקון", "אבוד"]


class ItemDialog(QDialog):
    def __init__(self, parent, db, item=None):
        super().__init__(parent)
        self.db = db
        self.item = item
        self.image_path = item['image_path'] if item and item['image_path'] else ""
        self.setWindowTitle("עריכת חפץ" if item else "הוספת חפץ חדש")
        self.setMinimumWidth(480)
        self.setLayoutDirection(Qt.RightToLeft)
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(24, 24, 24, 24)

        title = QLabel("✏️ עריכת חפץ" if self.item else "➕ הוספת חפץ חדש")
        title.setObjectName("pageTitle")
        layout.addWidget(title)
        layout.addWidget(sep_line())

        form = QFormLayout()
        form.setSpacing(12)
        form.setLabelAlignment(Qt.AlignRight)
        form.setFormAlignment(Qt.AlignRight | Qt.AlignTop)

        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("שם החפץ")
        form.addRow("שם החפץ *:", self.name_edit)

        self.category_cb = QComboBox()
        self.category_cb.addItems(CATEGORIES)
        form.addRow("קטגוריה:", self.category_cb)

        self.status_cb = QComboBox()
        self.status_cb.addItems(ITEM_STATUSES)
        form.addRow("סטטוס:", self.status_cb)

        self.serial_edit = QLineEdit()
        self.serial_edit.setPlaceholderText("אופציונלי")
        form.addRow("מספר סידורי:", self.serial_edit)

        self.location_edit = QLineEdit()
        self.location_edit.setPlaceholderText("מדף, ארון, חדר...")
        form.addRow("מיקום פיזי:", self.location_edit)

        self.notes_edit = QTextEdit()
        self.notes_edit.setMaximumHeight(80)
        self.notes_edit.setPlaceholderText("הערות על מצב החפץ...")
        form.addRow("הערות:", self.notes_edit)

        layout.addLayout(form)

        # Image
        img_layout = QHBoxLayout()
        self.img_label = QLabel("אין תמונה")
        self.img_label.setFixedSize(80, 80)
        self.img_label.setStyleSheet(
            f"background:{COLORS['background']}; border:1px solid {COLORS['border']}; border-radius:8px;"
        )
        self.img_label.setAlignment(Qt.AlignCenter)
        img_btn = QPushButton("📷 בחר תמונה")
        img_btn.setObjectName("secondaryBtn")
        img_btn.clicked.connect(self._pick_image)
        img_layout.addWidget(self.img_label)
        img_layout.addWidget(img_btn)
        img_layout.addStretch()
        layout.addLayout(img_layout)

        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        cancel = QPushButton("ביטול")
        cancel.setObjectName("secondaryBtn")
        cancel.clicked.connect(self.reject)
        save = QPushButton("💾 שמור")
        save.setObjectName("primaryBtn")
        save.clicked.connect(self._save)
        btn_layout.addWidget(cancel)
        btn_layout.addWidget(save)
        layout.addLayout(btn_layout)

        # Fill data
        if self.item:
            self.name_edit.setText(self.item['name'])
            idx = self.category_cb.findText(self.item['category'])
            if idx >= 0:
                self.category_cb.setCurrentIndex(idx)
            idx2 = self.status_cb.findText(self.item['status'])
            if idx2 >= 0:
                self.status_cb.setCurrentIndex(idx2)
            self.serial_edit.setText(self.item['serial_number'] or "")
            self.location_edit.setText(self.item['location'] or "")
            self.notes_edit.setPlainText(self.item['notes'] or "")
            if self.image_path and os.path.exists(self.image_path):
                pix = QPixmap(self.image_path).scaled(80, 80, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.img_label.setPixmap(pix)

    def _pick_image(self):
        path, _ = QFileDialog.getOpenFileName(self, "בחר תמונה", "", "תמונות (*.png *.jpg *.jpeg *.bmp)")
        if path:
            dest = str(IMAGES_DIR / Path(path).name)
            import shutil
            shutil.copy2(path, dest)
            self.image_path = dest
            pix = QPixmap(dest).scaled(80, 80, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.img_label.setPixmap(pix)

    def _save(self):
        name = self.name_edit.text().strip()
        if not name:
            QMessageBox.warning(self, "שגיאה", "יש להזין שם חפץ")
            return
        self.result_data = {
            "name": name,
            "category": self.category_cb.currentText(),
            "status": self.status_cb.currentText(),
            "serial": self.serial_edit.text().strip(),
            "location": self.location_edit.text().strip(),
            "notes": self.notes_edit.toPlainText().strip(),
            "image_path": self.image_path,
        }
        self.accept()


class BorrowerDialog(QDialog):
    def __init__(self, parent, db, borrower=None):
        super().__init__(parent)
        self.db = db
        self.borrower = borrower
        self.setWindowTitle("עריכת שואל" if borrower else "שואל חדש")
        self.setMinimumWidth(440)
        self.setLayoutDirection(Qt.RightToLeft)
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(24, 24, 24, 24)

        title = QLabel("✏️ עריכת שואל" if self.borrower else "👤 שואל חדש")
        title.setObjectName("pageTitle")
        layout.addWidget(title)
        layout.addWidget(sep_line())

        form = QFormLayout()
        form.setSpacing(12)
        form.setLabelAlignment(Qt.AlignRight)
        form.setFormAlignment(Qt.AlignRight | Qt.AlignTop)

        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("שם מלא")
        form.addRow("שם מלא *:", self.name_edit)

        self.phone_edit = QLineEdit()
        self.phone_edit.setPlaceholderText("05X-XXXXXXX")
        form.addRow("טלפון *:", self.phone_edit)

        self.email_edit = QLineEdit()
        self.email_edit.setPlaceholderText("אופציונלי")
        form.addRow("דוא״ל:", self.email_edit)

        self.address_edit = QLineEdit()
        form.addRow("כתובת:", self.address_edit)

        self.rating_spin = QSpinBox()
        self.rating_spin.setRange(1, 5)
        self.rating_spin.setValue(5)
        self.rating_spin.setPrefix("⭐ ")
        form.addRow("דירוג אמינות:", self.rating_spin)

        self.notes_edit = QTextEdit()
        self.notes_edit.setMaximumHeight(70)
        form.addRow("הערות:", self.notes_edit)

        layout.addLayout(form)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        cancel = QPushButton("ביטול")
        cancel.setObjectName("secondaryBtn")
        cancel.clicked.connect(self.reject)
        save = QPushButton("💾 שמור")
        save.setObjectName("primaryBtn")
        save.clicked.connect(self._save)
        btn_layout.addWidget(cancel)
        btn_layout.addWidget(save)
        layout.addLayout(btn_layout)

        if self.borrower:
            self.name_edit.setText(self.borrower['full_name'])
            self.phone_edit.setText(self.borrower['phone'])
            self.email_edit.setText(self.borrower['email'] or "")
            self.address_edit.setText(self.borrower['address'] or "")
            self.rating_spin.setValue(self.borrower['rating'] or 5)
            self.notes_edit.setPlainText(self.borrower['notes'] or "")

    def _save(self):
        name = self.name_edit.text().strip()
        phone = self.phone_edit.text().strip()
        if not name or not phone:
            QMessageBox.warning(self, "שגיאה", "יש להזין שם וטלפון")
            return
        self.result_data = {
            "full_name": name,
            "phone": phone,
            "email": self.email_edit.text().strip(),
            "address": self.address_edit.text().strip(),
            "rating": self.rating_spin.value(),
            "notes": self.notes_edit.toPlainText().strip(),
        }
        self.accept()


class LoanDialog(QDialog):
    def __init__(self, parent, db, preselect_item_id=None):
        super().__init__(parent)
        self.db = db
        self.preselect_item_id = preselect_item_id
        self.setWindowTitle("השאלה חדשה")
        self.setMinimumWidth(500)
        self.setLayoutDirection(Qt.RightToLeft)
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(24, 24, 24, 24)

        title = QLabel("📦 השאלה חדשה")
        title.setObjectName("pageTitle")
        layout.addWidget(title)
        layout.addWidget(sep_line())

        form = QFormLayout()
        form.setSpacing(12)
        form.setLabelAlignment(Qt.AlignRight)
        form.setFormAlignment(Qt.AlignRight | Qt.AlignTop)

        # Item search
        item_layout = QHBoxLayout()
        self.item_search = QLineEdit()
        self.item_search.setPlaceholderText("חפש חפץ לפי שם...")
        self.item_search.textChanged.connect(self._filter_items)
        self.item_cb = QComboBox()
        self.item_cb.setMinimumWidth(250)
        self._load_available_items()
        item_layout.addWidget(self.item_search)
        item_layout.addWidget(self.item_cb)
        form.addRow("חפץ *:", item_layout)

        # Borrower search
        bor_layout = QHBoxLayout()
        self.bor_search = QLineEdit()
        self.bor_search.setPlaceholderText("חפש שואל לפי שם/טלפון...")
        self.bor_search.textChanged.connect(self._filter_borrowers)
        self.bor_cb = QComboBox()
        self.bor_cb.setMinimumWidth(250)
        self._load_borrowers()
        new_bor_btn = QPushButton("+ חדש")
        new_bor_btn.setObjectName("secondaryBtn")
        new_bor_btn.clicked.connect(self._new_borrower)
        bor_layout.addWidget(self.bor_search)
        bor_layout.addWidget(self.bor_cb)
        bor_layout.addWidget(new_bor_btn)
        form.addRow("שואל *:", bor_layout)

        # Return date
        self.return_date = QDateEdit()
        self.return_date.setCalendarPopup(True)
        self.return_date.setDate(QDate.currentDate().addDays(7))
        self.return_date.setMinimumDate(QDate.currentDate())
        form.addRow("תאריך החזרה מתוכנן:", self.return_date)

        self.notes_edit = QLineEdit()
        self.notes_edit.setPlaceholderText("אופציונלי")
        form.addRow("הערות:", self.notes_edit)

        layout.addLayout(form)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        cancel = QPushButton("ביטול")
        cancel.setObjectName("secondaryBtn")
        cancel.clicked.connect(self.reject)
        save = QPushButton("✅ אשר השאלה")
        save.setObjectName("successBtn")
        save.clicked.connect(self._save)
        btn_layout.addWidget(cancel)
        btn_layout.addWidget(save)
        layout.addLayout(btn_layout)

    def _load_available_items(self, search=""):
        self.item_cb.clear()
        self._available_items = self.db.get_items(search=search, status="זמין")
        for it in self._available_items:
            self.item_cb.addItem(f"{it['name']} [{it['category']}]", it['id'])
        if self.preselect_item_id:
            for i, it in enumerate(self._available_items):
                if it['id'] == self.preselect_item_id:
                    self.item_cb.setCurrentIndex(i)

    def _filter_items(self, text):
        self._load_available_items(search=text)

    def _load_borrowers(self, search=""):
        self.bor_cb.clear()
        self._borrowers = self.db.get_borrowers(search=search)
        for b in self._borrowers:
            self.bor_cb.addItem(f"{b['full_name']} | {b['phone']}", b['id'])

    def _filter_borrowers(self, text):
        self._load_borrowers(search=text)

    def _new_borrower(self):
        dlg = BorrowerDialog(self, self.db)
        if dlg.exec() == QDialog.Accepted:
            d = dlg.result_data
            bid = self.db.add_borrower(d['full_name'], d['phone'], d['email'], d['address'], d['rating'], d['notes'])
            self._load_borrowers()
            for i in range(self.bor_cb.count()):
                if self.bor_cb.itemData(i) == bid:
                    self.bor_cb.setCurrentIndex(i)
                    break

    def _save(self):
        if self.item_cb.count() == 0:
            QMessageBox.warning(self, "שגיאה", "אין חפצים זמינים להשאלה")
            return
        if self.bor_cb.count() == 0:
            QMessageBox.warning(self, "שגיאה", "יש להוסיף שואל")
            return
        self.result_data = {
            "item_id": self.item_cb.currentData(),
            "borrower_id": self.bor_cb.currentData(),
            "planned_return": self.return_date.date().toString("yyyy-MM-dd"),
            "notes": self.notes_edit.text().strip(),
        }
        self.accept()


class EmailReminderDialog(QDialog):
    def __init__(self, parent, borrower_name, borrower_email, item_name, return_date):
        super().__init__(parent)
        self.setWindowTitle("שליחת תזכורת במייל")
        self.setMinimumWidth(500)
        self.setLayoutDirection(Qt.RightToLeft)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(14)

        title = QLabel("📧 שליחת תזכורת")
        title.setObjectName("pageTitle")
        layout.addWidget(title)
        layout.addWidget(sep_line())

        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignRight)
        form.setFormAlignment(Qt.AlignRight | Qt.AlignTop)

        self.smtp_server = QLineEdit("smtp.gmail.com")
        form.addRow("שרת SMTP:", self.smtp_server)

        self.smtp_port = QLineEdit("587")
        form.addRow("פורט:", self.smtp_port)

        self.sender_email = QLineEdit()
        self.sender_email.setPlaceholderText("your@gmail.com")
        form.addRow("מייל שולח:", self.sender_email)

        self.sender_pass = QLineEdit()
        self.sender_pass.setEchoMode(QLineEdit.Password)
        self.sender_pass.setPlaceholderText("סיסמה / App Password")
        form.addRow("סיסמה:", self.sender_pass)

        self.to_email = QLineEdit(borrower_email)
        form.addRow("מייל נמען:", self.to_email)

        body = f"""שלום {borrower_name},

תזכורת נעימה כי את/ה שואל/ת את הפריט: {item_name}
תאריך ההחזרה המתוכנן: {return_date}

אנא דאגו להחזיר את הפריט במועד.
בברכה,
מנהל הגמ"ח"""
        self.body_edit = QTextEdit()
        self.body_edit.setPlainText(body)
        self.body_edit.setMaximumHeight(160)
        form.addRow("תוכן ההודעה:", self.body_edit)

        layout.addLayout(form)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        cancel = QPushButton("ביטול")
        cancel.setObjectName("secondaryBtn")
        cancel.clicked.connect(self.reject)
        send = QPushButton("📧 שלח מייל")
        send.setObjectName("primaryBtn")
        send.clicked.connect(self._send)
        btn_layout.addWidget(cancel)
        btn_layout.addWidget(send)
        layout.addLayout(btn_layout)

    def _send(self):
        try:
            msg = MIMEMultipart()
            msg['From'] = self.sender_email.text()
            msg['To'] = self.to_email.text()
            msg['Subject'] = "תזכורת להחזרת פריט - גמ\"ח"
            msg.attach(MIMEText(self.body_edit.toPlainText(), 'plain', 'utf-8'))

            with smtplib.SMTP(self.smtp_server.text(), int(self.smtp_port.text())) as server:
                server.starttls()
                server.login(self.sender_email.text(), self.sender_pass.text())
                server.send_message(msg)

            QMessageBox.information(self, "הצלחה", "המייל נשלח בהצלחה!")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "שגיאה", f"שגיאה בשליחת המייל:\n{str(e)}")


# ─────────────────────────────────────────────
#  PAGE WIDGETS
# ─────────────────────────────────────────────

class DashboardBridge(QObject):
    """JS → Python bridge: lets the HTML dashboard trigger navigation/actions."""
    navigateRequested = Signal(int)

    @Slot(int)
    def navigate(self, idx):
        self.navigateRequested.emit(idx)


class DashboardPage(QWidget):
    navigate = Signal(int)

    def __init__(self, db):
        super().__init__()
        self.db = db
        self._use_web = WEBENGINE_AVAILABLE
        self._web_view = None
        self._bridge = None

    # ── public API expected by MainWindow ──────────────────────────
    def refresh(self):
        if self._use_web:
            self._render_web()
        else:
            self._build_classic()

    # ── HTML/CSS rich dashboard ─────────────────────────────────────
    def _ensure_web_view(self):
        if self._web_view is not None:
            return
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        self._web_view = QWebEngineView()
        self._web_view.setContextMenuPolicy(Qt.NoContextMenu)
        layout.addWidget(self._web_view)

        self._bridge = DashboardBridge()
        self._bridge.navigateRequested.connect(self.navigate.emit)
        self._channel = QWebChannel()
        self._channel.registerObject("bridge", self._bridge)
        self._web_view.page().setWebChannel(self._channel)

    def _render_web(self):
        try:
            self._ensure_web_view()
            html = self._build_html()
            self._web_view.setHtml(html, baseUrl=QUrl("qrc:/"))
        except Exception:
            # Any failure in the experimental web dashboard silently falls
            # back to the classic, fully-tested widget dashboard.
            self._use_web = False
            self._build_classic()

    def _build_html(self):
        stats = self.db.get_stats()
        overdue = self.db.get_overdue_loans()
        logs = self.db.get_audit_log(8)
        today_str = datetime.now().strftime("%A, %d/%m/%Y")

        total = max(stats['total'], 1)
        available = stats['available']
        loaned = stats['loaned']
        repair = stats['repair']
        overdue_n = stats['overdue']
        borrowers = stats['borrowers']

        pct_avail = round(available / total * 100)
        pct_loan = round(loaned / total * 100)
        pct_repair = round(repair / total * 100)

        # SVG ring chart geometry (single ring, three arcs)
        R = 70
        CIRC = 2 * 3.14159265 * R
        seg_avail = CIRC * (available / total)
        seg_loan = CIRC * (loaned / total)
        seg_repair = CIRC * (repair / total)
        seg_rest = max(CIRC - seg_avail - seg_loan - seg_repair, 0)

        # Overdue rows
        if overdue:
            today = date.today()
            rows_html = ""
            for loan in overdue[:6]:
                ret_d = date.fromisoformat(loan['planned_return'])
                days_late = (today - ret_d).days
                rows_html += f"""
                <tr>
                  <td class="item-name">{self._esc(loan['item_name'])}</td>
                  <td>{self._esc(loan['borrower_name'])}</td>
                  <td class="num">{self._esc(loan['phone'] or '')}</td>
                  <td class="num">{loan['planned_return']}</td>
                  <td><span class="late-pill">{days_late} ימים</span></td>
                </tr>"""
            overdue_section = f"""
            <section class="panel panel-overdue">
              <div class="panel-head">
                <h2>⚠️ חפצים באיחור <span class="count">{len(overdue)}</span></h2>
              </div>
              <table class="ledger">
                <thead><tr><th>חפץ</th><th>שואל</th><th>טלפון</th><th>תאריך החזרה</th><th>איחור</th></tr></thead>
                <tbody>{rows_html}</tbody>
              </table>
            </section>"""
        else:
            overdue_section = """
            <section class="panel panel-clean">
              <div class="clean-state">
                <div class="clean-icon">✓</div>
                <p>אין השאלות באיחור כעת — הכל מתנהל כשורה</p>
              </div>
            </section>"""

        # Activity log rows
        if logs:
            log_rows = ""
            for log in logs:
                log_rows += f"""
                <li class="log-row">
                  <span class="log-dot"></span>
                  <div class="log-body">
                    <span class="log-action">{self._esc(log['action'])}</span>
                    <span class="log-details">{self._esc(log['details'] or '')}</span>
                  </div>
                  <span class="log-time">{self._esc(log['timestamp'][:16])}</span>
                </li>"""
        else:
            log_rows = '<li class="log-empty">אין פעולות עדיין</li>'

        html = f"""<!DOCTYPE html>
<html lang="he" dir="rtl">
<head>
<meta charset="UTF-8">
<style>
{self._css()}
</style>
</head>
<body>
<div class="wrap">

  <header class="hero">
    <div class="hero-text">
      <p class="hero-eyebrow">{today_str}</p>
      <h1>שלום, ברוך הבא לגמ"ח</h1>
      <p class="hero-sub">תמונת מצב מהירה של החפצים, ההשאלות, ומה דורש את תשומת לבך היום</p>
    </div>
    <div class="hero-actions">
      <button class="btn btn-gold" onclick="go(8)">➕ השאלה חדשה</button>
      <button class="btn btn-outline" onclick="go(3)">✅ החזרת חפץ</button>
    </div>
  </header>

  <section class="overview">
    <div class="kpi-grid">
      <div class="kpi kpi-clickable" onclick="go(1)">
        <span class="kpi-icon">📦</span>
        <span class="kpi-val">{stats['total']}</span>
        <span class="kpi-label">סך הכל חפצים</span>
      </div>
      <div class="kpi kpi-clickable" onclick="go(1)">
        <span class="kpi-icon">✅</span>
        <span class="kpi-val accent-green">{available}</span>
        <span class="kpi-label">חפצים זמינים</span>
      </div>
      <div class="kpi kpi-clickable" onclick="go(3)">
        <span class="kpi-icon">📤</span>
        <span class="kpi-val accent-navy">{loaned}</span>
        <span class="kpi-label">מושאלים כעת</span>
      </div>
      <div class="kpi kpi-clickable {'kpi-alert' if overdue_n else ''}" onclick="go(3)">
        <span class="kpi-icon">⚠️</span>
        <span class="kpi-val accent-red">{overdue_n}</span>
        <span class="kpi-label">באיחור</span>
      </div>
      <div class="kpi kpi-clickable" onclick="go(1)">
        <span class="kpi-icon">🔧</span>
        <span class="kpi-val accent-gold">{repair}</span>
        <span class="kpi-label">בתיקון</span>
      </div>
      <div class="kpi kpi-clickable" onclick="go(2)">
        <span class="kpi-icon">👥</span>
        <span class="kpi-val">{borrowers}</span>
        <span class="kpi-label">שואלים רשומים</span>
      </div>
    </div>

    <div class="ring-card">
      <svg viewBox="0 0 180 180" class="ring-svg">
        <circle cx="90" cy="90" r="{R}" class="ring-bg"/>
        <circle cx="90" cy="90" r="{R}" class="ring-seg ring-avail"
          stroke-dasharray="{seg_avail} {CIRC - seg_avail}" stroke-dashoffset="0"/>
        <circle cx="90" cy="90" r="{R}" class="ring-seg ring-loan"
          stroke-dasharray="{seg_loan} {CIRC - seg_loan}" stroke-dashoffset="-{seg_avail}"/>
        <circle cx="90" cy="90" r="{R}" class="ring-seg ring-repair"
          stroke-dasharray="{seg_repair} {CIRC - seg_repair}" stroke-dashoffset="-{seg_avail + seg_loan}"/>
        <text x="90" y="84" class="ring-num">{stats['total']}</text>
        <text x="90" y="106" class="ring-label">חפצים</text>
      </svg>
      <div class="ring-legend">
        <div class="legend-row"><span class="dot dot-avail"></span>זמינים <b>{available}</b> <span class="pct">({pct_avail}%)</span></div>
        <div class="legend-row"><span class="dot dot-loan"></span>מושאלים <b>{loaned}</b> <span class="pct">({pct_loan}%)</span></div>
        <div class="legend-row"><span class="dot dot-repair"></span>בתיקון <b>{repair}</b> <span class="pct">({pct_repair}%)</span></div>
      </div>
    </div>
  </section>

  {overdue_section}

  <section class="panel panel-activity">
    <div class="panel-head"><h2>📋 פעולות אחרונות</h2></div>
    <ul class="log-list">
      {log_rows}
    </ul>
  </section>

</div>

<script src="qrc:///qtwebchannel/qwebchannel.js"></script>
<script>
  let bridge = null;
  new QWebChannel(qt.webChannelTransport, function(channel) {{
    bridge = channel.objects.bridge;
  }});
  function go(idx) {{
    if (bridge) {{ bridge.navigate(idx); }}
  }}
</script>
</body>
</html>"""
        return html

    @staticmethod
    def _esc(s):
        s = str(s) if s is not None else ""
        return (s.replace("&", "&amp;").replace("<", "&lt;")
                 .replace(">", "&gt;").replace('"', "&quot;"))

    @staticmethod
    def _css():
        root_vars = f"""
        :root {{
          --navy: {COLORS.get('html_navy', '#0F1B2E')};
          --navy-2: {COLORS.get('html_navy2', '#16263F')};
          --gold: {COLORS.get('html_gold', '#D4A24C')};
          --gold-light: {COLORS.get('html_gold_light', '#E8C588')};
          --sage: {COLORS.get('html_sage', '#5B8C6E')};
          --sage-light: {COLORS.get('html_sage_light', '#E3EEE6')};
          --rose: {COLORS.get('html_rose', '#C75D4D')};
          --rose-light: {COLORS.get('html_rose_light', '#F7E2DE')};
          --parchment: {COLORS.get('html_parchment', '#F7F3EC')};
          --paper: {COLORS.get('card', '#FFFFFF')};
          --ink: {COLORS.get('text_primary', '#1F2A33')};
          --ink-soft: {COLORS.get('text_secondary', '#6B7780')};
          --line: {COLORS.get('html_line', '#E7E0D2')};
        }}
        """
        rest = """
        * { box-sizing: border-box; }
        body {
          margin: 0;
          background: var(--parchment);
          font-family: "Segoe UI", "Segoe UI Semibold", Tahoma, Arial, sans-serif;
          color: var(--ink);
        }
        .wrap { padding: 26px 30px 40px; max-width: 1180px; margin: 0 auto; }

        /* Hero */
        .hero {
          background: linear-gradient(135deg, var(--navy) 0%, var(--navy-2) 60%, #1d3450 100%);
          border-radius: 18px;
          padding: 28px 32px;
          display: flex;
          align-items: center;
          justify-content: space-between;
          gap: 24px;
          box-shadow: 0 10px 30px rgba(15,27,46,0.25);
          position: relative;
          overflow: hidden;
          animation: fadeUp .5s ease both;
        }
        .hero::after {
          content: "";
          position: absolute; inset: 0;
          background: radial-gradient(circle at 85% 20%, rgba(212,162,76,0.18), transparent 55%);
          pointer-events: none;
        }
        .hero-eyebrow {
          color: var(--gold-light);
          font-size: 12.5px;
          letter-spacing: .04em;
          margin: 0 0 6px;
          font-weight: 600;
        }
        .hero h1 {
          color: #fff;
          font-size: 26px;
          margin: 0 0 8px;
          font-weight: 700;
        }
        .hero-sub {
          color: #B9C4D4;
          font-size: 13.5px;
          margin: 0;
          max-width: 480px;
        }
        .hero-actions { display: flex; gap: 10px; flex-shrink: 0; z-index: 1; }
        .btn {
          border: none;
          border-radius: 10px;
          padding: 12px 18px;
          font-size: 13.5px;
          font-weight: 600;
          cursor: pointer;
          font-family: inherit;
          transition: transform .15s ease, box-shadow .15s ease, filter .15s ease;
        }
        .btn:hover { transform: translateY(-2px); }
        .btn-gold {
          background: linear-gradient(135deg, var(--gold), #c08e3a);
          color: #1a1206;
          box-shadow: 0 6px 16px rgba(212,162,76,0.35);
        }
        .btn-gold:hover { filter: brightness(1.06); }
        .btn-outline {
          background: rgba(255,255,255,0.07);
          color: #EAF0F7;
          border: 1px solid rgba(255,255,255,0.25);
        }
        .btn-outline:hover { background: rgba(255,255,255,0.14); }

        /* Overview: ring + kpis */
        .overview {
          display: grid;
          grid-template-columns: 1fr 250px;
          gap: 18px;
          margin-top: 20px;
          animation: fadeUp .55s ease both;
          animation-delay: .05s;
        }
        .ring-card {
          background: var(--paper);
          border: 1px solid var(--line);
          border-radius: 16px;
          padding: 18px;
          display: flex;
          flex-direction: column;
          align-items: center;
          gap: 12px;
          box-shadow: 0 2px 10px rgba(15,27,46,0.04);
        }
        .ring-svg { width: 150px; height: 150px; transform: rotate(-90deg); }
        .ring-bg { fill: none; stroke: var(--line); stroke-width: 14; }
        .ring-seg { fill: none; stroke-width: 14; stroke-linecap: round;
          transition: stroke-dasharray 1s ease; }
        .ring-avail { stroke: var(--sage); }
        .ring-loan { stroke: var(--navy); }
        .ring-repair { stroke: var(--gold); }
        .ring-num, .ring-label {
          transform: rotate(90deg);
          transform-origin: 90px 90px;
          text-anchor: middle;
          font-family: inherit;
        }
        .ring-num { font-size: 30px; font-weight: 700; fill: var(--ink); }
        .ring-label { font-size: 11px; fill: var(--ink-soft); }
        .ring-legend { width: 100%; display: flex; flex-direction: column; gap: 7px; }
        .legend-row { font-size: 12.5px; display: flex; align-items: center; gap: 6px; color: var(--ink-soft); }
        .legend-row b { color: var(--ink); margin: 0 2px; }
        .pct { color: #A8B0B8; font-size: 11px; }
        .dot { width: 9px; height: 9px; border-radius: 50%; display: inline-block; flex-shrink: 0; }
        .dot-avail { background: var(--sage); }
        .dot-loan { background: var(--navy); }
        .dot-repair { background: var(--gold); }

        .kpi-grid {
          display: grid;
          grid-template-columns: repeat(3, 1fr);
          gap: 12px;
        }
        .kpi {
          background: var(--paper);
          border: 1px solid var(--line);
          border-radius: 14px;
          padding: 16px 14px;
          display: flex;
          flex-direction: column;
          gap: 4px;
          box-shadow: 0 2px 10px rgba(15,27,46,0.04);
          transition: transform .15s ease, box-shadow .15s ease, border-color .15s ease;
        }
        .kpi-clickable { cursor: pointer; }
        .kpi-clickable:hover {
          transform: translateY(-3px);
          box-shadow: 0 10px 22px rgba(15,27,46,0.10);
          border-color: #DCD3BC;
        }
        .kpi-icon { font-size: 19px; }
        .kpi-val { font-size: 25px; font-weight: 700; color: var(--ink); }
        .kpi-label { font-size: 12px; color: var(--ink-soft); }
        .accent-green { color: var(--sage); }
        .accent-navy { color: var(--navy); }
        .accent-red { color: var(--rose); }
        .accent-gold { color: #B9883A; }
        .kpi-alert { border-color: var(--rose); background: linear-gradient(180deg, var(--paper), var(--rose-light)); }

        /* Panels */
        .panel {
          background: var(--paper);
          border: 1px solid var(--line);
          border-radius: 16px;
          margin-top: 18px;
          padding: 18px 20px;
          box-shadow: 0 2px 10px rgba(15,27,46,0.04);
          animation: fadeUp .6s ease both;
          animation-delay: .1s;
        }
        .panel-head { display: flex; align-items: center; gap: 10px; margin-bottom: 12px; }
        .panel-head h2 { font-size: 15.5px; margin: 0; font-weight: 700; display:flex; align-items:center; gap:8px;}
        .panel-overdue .panel-head h2 { color: var(--rose); }
        .count {
          background: var(--rose); color: #fff; border-radius: 999px;
          font-size: 11px; padding: 2px 9px; font-weight: 700;
        }
        .ledger { width: 100%; border-collapse: collapse; font-size: 13px; }
        .ledger th {
          text-align: right; color: var(--ink-soft); font-weight: 600;
          font-size: 11.5px; padding: 6px 8px; border-bottom: 1px solid var(--line);
        }
        .ledger td { padding: 9px 8px; border-bottom: 1px solid var(--line); }
        .ledger tr:hover td { background: var(--parchment); }
        .item-name { font-weight: 600; }
        .num { color: var(--ink-soft); }
        .late-pill {
          background: var(--rose-light); color: var(--rose);
          padding: 3px 10px; border-radius: 999px; font-size: 11.5px; font-weight: 700;
        }

        .clean-state { text-align: center; padding: 20px 0; color: var(--ink-soft); }
        .clean-icon {
          width: 42px; height: 42px; border-radius: 50%;
          background: var(--sage-light); color: var(--sage);
          display: flex; align-items: center; justify-content: center;
          font-size: 20px; margin: 0 auto 10px; font-weight: 700;
        }

        .log-list { list-style: none; margin: 0; padding: 0; display: flex; flex-direction: column; }
        .log-row {
          display: flex; align-items: center; gap: 10px;
          padding: 9px 4px; border-bottom: 1px solid var(--line);
        }
        .log-row:last-child { border-bottom: none; }
        .log-dot { width: 7px; height: 7px; border-radius: 50%; background: var(--gold); flex-shrink: 0; }
        .log-body { display: flex; flex-direction: column; flex: 1; min-width: 0; }
        .log-action { font-size: 13px; font-weight: 600; }
        .log-details { font-size: 11.5px; color: var(--ink-soft); white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
        .log-time { font-size: 11px; color: #A8B0B8; flex-shrink: 0; }
        .log-empty { color: var(--ink-soft); font-size: 13px; padding: 10px 4px; }

        @keyframes fadeUp {
          from { opacity: 0; transform: translateY(10px); }
          to { opacity: 1; transform: translateY(0); }
        }
        """
        return root_vars + rest

    # ── classic Qt-widgets fallback (unchanged behaviour) ───────────
    def _build_classic(self):
        old_layout = self.layout()
        if old_layout is not None:
            while old_layout.count():
                item = old_layout.takeAt(0)
                w = item.widget()
                if w:
                    w.setParent(None)
                    w.deleteLater()
            import shiboken6
            shiboken6.delete(old_layout)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(20)

        title_row = QHBoxLayout()
        title = QLabel("🏠 לוח בקרה")
        title.setObjectName("pageTitle")
        time_lbl = QLabel(datetime.now().strftime("%A, %d/%m/%Y"))
        time_lbl.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 13px;")
        title_row.addWidget(title)
        title_row.addStretch()
        title_row.addWidget(time_lbl)
        layout.addLayout(title_row)

        quick = QHBoxLayout()
        quick.setSpacing(12)
        new_loan_btn = QPushButton("➕  השאלה חדשה")
        new_loan_btn.setObjectName("successBtn")
        new_loan_btn.setFixedHeight(42)
        new_loan_btn.clicked.connect(lambda: self.navigate.emit(8))
        return_btn = QPushButton("✅  החזרת חפץ")
        return_btn.setObjectName("primaryBtn")
        return_btn.setFixedHeight(42)
        return_btn.clicked.connect(lambda: self.navigate.emit(3))
        quick.addWidget(new_loan_btn)
        quick.addWidget(return_btn)
        quick.addStretch()
        layout.addLayout(quick)

        stats_grid = QHBoxLayout()
        stats_grid.setSpacing(14)

        stats = self.db.get_stats()
        cards_data = [
            ("סך הכל חפצים", stats['total'], COLORS['primary_light'], "📦"),
            ("חפצים זמינים", stats['available'], COLORS['accent'], "✅"),
            ("מושאלים כעת", stats['loaned'], COLORS['primary'], "📤"),
            ("באיחור", stats['overdue'], COLORS['danger'], "⚠️"),
            ("בתיקון", stats['repair'], COLORS['warning'], "🔧"),
            ("שואלים רשומים", stats['borrowers'], COLORS['text_secondary'], "👥"),
        ]
        self._stat_vals = {}
        for label, val, color, icon in cards_data:
            frame, val_lbl = stat_card(label, val, color, icon)
            self._stat_vals[label] = val_lbl
            stats_grid.addWidget(frame)
        layout.addLayout(stats_grid)

        overdue = self.db.get_overdue_loans()
        if overdue:
            overdue_frame = card_frame()
            overdue_layout = QVBoxLayout(overdue_frame)
            overdue_layout.setContentsMargins(16, 14, 16, 14)

            h = QHBoxLayout()
            t = QLabel(f"⚠️  חפצים באיחור ({len(overdue)})")
            t.setObjectName("sectionTitle")
            t.setStyleSheet(f"color: {COLORS['danger']}; font-size: 15px; font-weight: 700;")
            h.addWidget(t)
            h.addStretch()
            overdue_layout.addLayout(h)

            tbl = QTableWidget(len(overdue), 5)
            rtl_table(tbl)
            tbl.setHorizontalHeaderLabels(["חפץ", "שואל", "טלפון", "תאריך החזרה", "ימי איחור"])
            tbl.horizontalHeader().setStretchLastSection(False)
            tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
            tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
            tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
            tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
            tbl.verticalHeader().setVisible(False)
            tbl.setMaximumHeight(200)

            today = date.today()
            for row, loan in enumerate(overdue):
                ret_d = date.fromisoformat(loan['planned_return'])
                days_late = (today - ret_d).days
                items_data = [
                    loan['item_name'], loan['borrower_name'],
                    loan['phone'], loan['planned_return'],
                    f"{days_late} ימים"
                ]
                for col, val in enumerate(items_data):
                    item = QTableWidgetItem(str(val))
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    if col == 4:
                        item.setForeground(QColor(COLORS['danger']))
                    tbl.setItem(row, col, item)

            tbl.setFixedHeight(min(200, 44 + len(overdue) * 44))
            overdue_layout.addWidget(tbl)
            layout.addWidget(overdue_frame)

        recent_frame = card_frame()
        recent_layout = QVBoxLayout(recent_frame)
        recent_layout.setContentsMargins(16, 14, 16, 14)
        t2 = QLabel("📋  פעולות אחרונות")
        t2.setObjectName("sectionTitle")
        recent_layout.addWidget(t2)

        logs = self.db.get_audit_log(10)
        if logs:
            log_tbl = QTableWidget(len(logs), 3)
            rtl_table(log_tbl)
            log_tbl.setHorizontalHeaderLabels(["פעולה", "פרטים", "זמן"])
            log_tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
            log_tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
            log_tbl.verticalHeader().setVisible(False)
            log_tbl.setMaximumHeight(220)
            for row, log in enumerate(logs):
                for col, val in enumerate([log['action'], log['details'], log['timestamp'][:16]]):
                    itm = QTableWidgetItem(str(val))
                    itm.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    log_tbl.setItem(row, col, itm)
            recent_layout.addWidget(log_tbl)
        else:
            recent_layout.addWidget(QLabel("אין פעולות עדיין"))

        layout.addWidget(recent_frame)
        layout.addStretch()

class ItemsPage(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        # Header
        h = QHBoxLayout()
        title = QLabel("📦 ניהול חפצים")
        title.setObjectName("pageTitle")
        h.addWidget(title)
        h.addStretch()
        add_cat_btn = QPushButton("🏷️  הוסף קטגוריה")
        add_cat_btn.setObjectName("secondaryBtn")
        add_cat_btn.clicked.connect(self._add_category)
        h.addWidget(add_cat_btn)
        add_btn = QPushButton("➕  הוספת חפץ")
        add_btn.setObjectName("successBtn")
        add_btn.clicked.connect(self._add_item)
        h.addWidget(add_btn)
        layout.addLayout(h)

        # Filters
        filter_frame = card_frame()
        fl = QHBoxLayout(filter_frame)
        fl.setContentsMargins(14, 10, 14, 10)

        self.search_box = QLineEdit()
        self.search_box.setObjectName("searchBox")
        self.search_box.setPlaceholderText("🔍  חיפוש חפץ...")
        self.search_box.setFixedWidth(220)
        self.search_box.textChanged.connect(self._refresh_table)

        self.cat_filter = QComboBox()
        self.cat_filter.addItem("כל הקטגוריות", "")
        for c in CATEGORIES:
            self.cat_filter.addItem(c, c)
        self.cat_filter.currentIndexChanged.connect(self._refresh_table)

        self.status_filter = QComboBox()
        self.status_filter.addItem("כל הסטטוסים", "")
        for s in ITEM_STATUSES:
            self.status_filter.addItem(s, s)
        self.status_filter.currentIndexChanged.connect(self._refresh_table)

        fl.addWidget(QLabel("סינון:"))
        fl.addWidget(self.cat_filter)
        fl.addWidget(self.status_filter)
        fl.addStretch()
        fl.addWidget(self.search_box)
        layout.addWidget(filter_frame)

        # Table
        self.table = QTableWidget(0, 7)
        rtl_table(self.table)
        self.table.setHorizontalHeaderLabels(["שם החפץ", "קטגוריה", "סטטוס", "מיקום", "מספר סידורי", "הערות", "פעולות"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.Stretch)
        self.table.horizontalHeader().setMinimumSectionSize(90)
        self.table.setColumnWidth(1, 110)
        self.table.setColumnWidth(6, 90)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet(self.table.styleSheet() + "QTableWidget {alternate-background-color: #F7FAFC;}")
        layout.addWidget(self.table)

        self._refresh_table()

    def _refresh_table(self):
        search = self.search_box.text()
        cat = self.cat_filter.currentData()
        status = self.status_filter.currentData()
        items = self.db.get_items(search=search, category=cat, status=status)

        self.table.setRowCount(len(items))
        for row, it in enumerate(items):
            self.table.setRowHeight(row, 46)

            for col, val in enumerate([
                it['name'], it['category'], it['status'],
                it['location'] or "", it['serial_number'] or "", it['notes'] or ""
            ]):
                cell = QTableWidgetItem(str(val))
                cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if col == 2:  # status
                    color_map = {
                        "זמין": COLORS['accent'], "מושאל": COLORS['primary_light'],
                        "בתיקון": COLORS['warning'], "אבוד": COLORS['danger']
                    }
                    cell.setForeground(QColor(color_map.get(val, COLORS['text_primary'])))
                self.table.setItem(row, col, cell)

            # Action buttons
            action_widget = QWidget()
            action_widget.setLayoutDirection(Qt.RightToLeft)
            action_layout = QHBoxLayout(action_widget)
            action_layout.setContentsMargins(4, 2, 4, 2)
            action_layout.setSpacing(4)

            edit_btn = QPushButton("✏️")
            edit_btn.setObjectName("iconBtn")
            edit_btn.setToolTip("עריכה")
            edit_btn.clicked.connect(lambda _, i=it['id']: self._edit_item(i))

            del_btn = QPushButton("🗑️")
            del_btn.setObjectName("iconBtn")
            del_btn.setToolTip("מחיקה")
            del_btn.clicked.connect(lambda _, i=it['id']: self._delete_item(i))

            action_layout.addWidget(edit_btn)
            action_layout.addWidget(del_btn)
            self.table.setCellWidget(row, 6, action_widget)

    def _add_item(self):
        dlg = ItemDialog(self, self.db)
        if dlg.exec() == QDialog.Accepted:
            d = dlg.result_data
            self.db.add_item(d['name'], d['category'], d['status'], d['serial'], d['location'], d['notes'], d['image_path'])
            self._refresh_table()

    def _add_category(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("הוספת קטגוריה חדשה")
        dlg.setLayoutDirection(Qt.RightToLeft)
        dlg.setFixedWidth(320)
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)
        lbl = QLabel("שם הקטגוריה החדשה:")
        lbl.setObjectName("sectionTitle")
        layout.addWidget(lbl)
        name_edit = QLineEdit()
        name_edit.setPlaceholderText("לדוגמה: ספורט, ילדים...")
        layout.addWidget(name_edit)
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        cancel_btn = QPushButton("ביטול")
        cancel_btn.setObjectName("secondaryBtn")
        cancel_btn.clicked.connect(dlg.reject)
        ok_btn = QPushButton("✅ הוסף")
        ok_btn.setObjectName("primaryBtn")
        def _confirm():
            name = name_edit.text().strip()
            if not name:
                QMessageBox.warning(dlg, "שגיאה", "יש להזין שם קטגוריה")
                return
            if name in CATEGORIES:
                QMessageBox.warning(dlg, "שגיאה", "קטגוריה זו כבר קיימת")
                return
            CATEGORIES.append(name)
            # Update filter combo
            self.cat_filter.addItem(name, name)
            dlg.accept()
            QMessageBox.information(self, "הצלחה", f"הקטגוריה '{name}' נוספה!")
        ok_btn.clicked.connect(_confirm)
        name_edit.returnPressed.connect(_confirm)
        btn_row.addWidget(cancel_btn)
        btn_row.addWidget(ok_btn)
        layout.addLayout(btn_row)
        dlg.exec()

    def _edit_item(self, item_id):
        item = self.db.get_item(item_id)
        dlg = ItemDialog(self, self.db, item)
        if dlg.exec() == QDialog.Accepted:
            d = dlg.result_data
            self.db.update_item(item_id, d['name'], d['category'], d['status'], d['serial'], d['location'], d['notes'], d['image_path'])
            self._refresh_table()

    def _delete_item(self, item_id):
        reply = QMessageBox.question(self, "אישור מחיקה", "האם למחוק חפץ זה?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            if not self.db.delete_item(item_id):
                QMessageBox.warning(self, "שגיאה", "לא ניתן למחוק חפץ שמושאל כרגע")
            else:
                self._refresh_table()

    def refresh(self):
        self._refresh_table()


class BorrowersPage(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        h = QHBoxLayout()
        title = QLabel("👥 ניהול שואלים")
        title.setObjectName("pageTitle")
        h.addWidget(title)
        h.addStretch()
        add_btn = QPushButton("➕  שואל חדש")
        add_btn.setObjectName("successBtn")
        add_btn.clicked.connect(self._add_borrower)
        h.addWidget(add_btn)
        layout.addLayout(h)

        # Search
        filter_frame = card_frame()
        fl = QHBoxLayout(filter_frame)
        fl.setContentsMargins(14, 10, 14, 10)
        self.search_box = QLineEdit()
        self.search_box.setObjectName("searchBox")
        self.search_box.setPlaceholderText("🔍  חיפוש לפי שם או טלפון...")
        self.search_box.setFixedWidth(260)
        self.search_box.textChanged.connect(self._refresh_table)
        fl.addStretch()
        fl.addWidget(self.search_box)
        layout.addWidget(filter_frame)

        self.table = QTableWidget(0, 7)
        rtl_table(self.table)
        self.table.setHorizontalHeaderLabels(["שם מלא", "טלפון", "דוא״ל", "כתובת", "דירוג", "פעיל", "פעולות"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet(self.table.styleSheet() + "QTableWidget {alternate-background-color: #F7FAFC;}")
        layout.addWidget(self.table)
        self._refresh_table()

    def _refresh_table(self):
        borrowers = self.db.get_borrowers(search=self.search_box.text())
        self.table.setRowCount(len(borrowers))
        for row, b in enumerate(borrowers):
            self.table.setRowHeight(row, 46)
            # Active loans count
            active = self.db.conn.execute(
                "SELECT COUNT(*) FROM loans WHERE borrower_id=? AND status IN ('פעיל','באיחור')", (b['id'],)
            ).fetchone()[0]
            stars = "⭐" * (b['rating'] or 0)
            for col, val in enumerate([
                b['full_name'], b['phone'], b['email'] or "",
                b['address'] or "", stars, str(active)
            ]):
                cell = QTableWidgetItem(val)
                cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if col == 5 and active > 0:
                    cell.setForeground(QColor(COLORS['primary_light']))
                self.table.setItem(row, col, cell)

            # Actions
            aw = QWidget()
            aw.setLayoutDirection(Qt.RightToLeft)
            al = QHBoxLayout(aw)
            al.setContentsMargins(4, 2, 4, 2)
            al.setSpacing(4)
            edit_btn = QPushButton("✏️")
            edit_btn.setObjectName("iconBtn")
            edit_btn.clicked.connect(lambda _, i=b['id']: self._edit_borrower(i))
            del_btn = QPushButton("🗑️")
            del_btn.setObjectName("iconBtn")
            del_btn.clicked.connect(lambda _, i=b['id']: self._delete_borrower(i))
            al.addWidget(edit_btn)
            al.addWidget(del_btn)
            self.table.setCellWidget(row, 6, aw)

    def _add_borrower(self):
        dlg = BorrowerDialog(self, self.db)
        if dlg.exec() == QDialog.Accepted:
            d = dlg.result_data
            self.db.add_borrower(d['full_name'], d['phone'], d['email'], d['address'], d['rating'], d['notes'])
            self._refresh_table()

    def _edit_borrower(self, bid):
        b = self.db.get_borrower(bid)
        dlg = BorrowerDialog(self, self.db, b)
        if dlg.exec() == QDialog.Accepted:
            d = dlg.result_data
            self.db.update_borrower(bid, d['full_name'], d['phone'], d['email'], d['address'], d['rating'], d['notes'])
            self._refresh_table()

    def _delete_borrower(self, bid):
        reply = QMessageBox.question(self, "אישור", "למחוק שואל זה?", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            if not self.db.delete_borrower(bid):
                QMessageBox.warning(self, "שגיאה", "לא ניתן למחוק שואל עם השאלות פעילות")
            else:
                self._refresh_table()

    def refresh(self):
        self._refresh_table()


class LoansPage(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        h = QHBoxLayout()
        title = QLabel("📋 ניהול השאלות")
        title.setObjectName("pageTitle")
        h.addWidget(title)
        h.addStretch()
        layout.addLayout(h)

        tabs = QTabWidget()
        tabs.addTab(self._build_active_tab(), "📤  השאלות פעילות")
        tabs.addTab(self._build_overdue_tab(), "⚠️  באיחור")
        tabs.addTab(self._build_history_tab(), "📜  היסטוריה")
        layout.addWidget(tabs)

    def _build_active_tab(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(12, 12, 12, 12)

        self.active_table = QTableWidget(0, 7)
        rtl_table(self.active_table)
        self.active_table.setHorizontalHeaderLabels(
            ["חפץ", "קטגוריה", "שואל", "טלפון", "תאריך השאלה", "להחזרה עד", "פעולות"]
        )
        self.active_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.active_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.active_table.horizontalHeader().setMinimumSectionSize(90)
        self.active_table.setColumnWidth(6, 170)
        self.active_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.active_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.active_table.verticalHeader().setVisible(False)
        layout.addWidget(self.active_table)
        self._refresh_active()
        return w

    def _refresh_active(self):
        loans = self.db.get_loans(active_only=True)
        self.active_table.setRowCount(len(loans))
        for row, loan in enumerate(loans):
            self.active_table.setRowHeight(row, 46)
            loan_date = loan['loan_date'][:10] if loan['loan_date'] else ""
            for col, val in enumerate([
                loan['item_name'], loan['category'], loan['borrower_name'],
                loan['phone'], loan_date, loan['planned_return']
            ]):
                cell = QTableWidgetItem(str(val))
                cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.active_table.setItem(row, col, cell)

            aw = QWidget()
            aw.setLayoutDirection(Qt.RightToLeft)
            al = QHBoxLayout(aw)
            al.setContentsMargins(4, 2, 4, 2)
            al.setSpacing(4)

            ret_btn = QPushButton("✅ החזר")
            ret_btn.setObjectName("successBtn")
            ret_btn.setFixedHeight(30)
            ret_btn.setMinimumWidth(90)
            ret_btn.clicked.connect(lambda _, lid=loan['id'], iid=loan['item_id']: self._return_item(lid, iid))

            email_btn = QPushButton("📧")
            email_btn.setObjectName("iconBtn")
            email_btn.setToolTip("שלח תזכורת מייל")
            email_btn.clicked.connect(lambda _, l=loan: self._send_reminder(l))

            al.addWidget(ret_btn)
            al.addWidget(email_btn)
            self.active_table.setCellWidget(row, 6, aw)

    def _build_overdue_tab(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(12, 12, 12, 12)

        self.overdue_table = QTableWidget(0, 6)
        rtl_table(self.overdue_table)
        self.overdue_table.setHorizontalHeaderLabels(
            ["חפץ", "שואל", "טלפון", "תאריך החזרה", "ימי איחור", "פעולות"]
        )
        self.overdue_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.overdue_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.overdue_table.horizontalHeader().setMinimumSectionSize(90)
        self.overdue_table.setColumnWidth(5, 170)
        self.overdue_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.overdue_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.overdue_table.verticalHeader().setVisible(False)
        layout.addWidget(self.overdue_table)
        self._refresh_overdue()
        return w

    def _refresh_overdue(self):
        loans = self.db.get_overdue_loans()
        today = date.today()
        self.overdue_table.setRowCount(len(loans))
        for row, loan in enumerate(loans):
            self.overdue_table.setRowHeight(row, 46)
            ret_d = date.fromisoformat(loan['planned_return'])
            days_late = (today - ret_d).days
            for col, val in enumerate([
                loan['item_name'], loan['borrower_name'],
                loan['phone'], loan['planned_return'], f"{days_late} ימים"
            ]):
                cell = QTableWidgetItem(str(val))
                cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if col == 4:
                    cell.setForeground(QColor(COLORS['danger']))
                self.overdue_table.setItem(row, col, cell)

            aw = QWidget()
            aw.setLayoutDirection(Qt.RightToLeft)
            al = QHBoxLayout(aw)
            al.setContentsMargins(4, 2, 4, 2)
            ret_btn = QPushButton("✅ החזר")
            ret_btn.setObjectName("dangerBtn")
            ret_btn.setFixedHeight(30)
            ret_btn.setMinimumWidth(90)
            ret_btn.clicked.connect(lambda _, lid=loan['id'], iid=loan['item_id']: self._return_item(lid, iid))
            email_btn = QPushButton("📧")
            email_btn.setObjectName("iconBtn")
            email_btn.clicked.connect(lambda _, l=loan: self._send_reminder(l))
            al.addWidget(ret_btn)
            al.addWidget(email_btn)
            self.overdue_table.setCellWidget(row, 5, aw)

    def _build_history_tab(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(12, 12, 12, 12)

        self.history_table = QTableWidget(0, 6)
        rtl_table(self.history_table)
        self.history_table.setHorizontalHeaderLabels(
            ["חפץ", "שואל", "תאריך השאלה", "תאריך החזרה מתוכנן", "הוחזר בפועל", "סטטוס"]
        )
        self.history_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.history_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.history_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.history_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.history_table.verticalHeader().setVisible(False)
        layout.addWidget(self.history_table)
        self._refresh_history()
        return w

    def _refresh_history(self):
        loans = self.db.get_loans()
        self.history_table.setRowCount(len(loans))
        for row, loan in enumerate(loans):
            self.history_table.setRowHeight(row, 44)
            for col, val in enumerate([
                loan['item_name'], loan['borrower_name'],
                loan['loan_date'][:10] if loan['loan_date'] else "",
                loan['planned_return'],
                loan['actual_return'][:10] if loan['actual_return'] else "—",
                loan['status']
            ]):
                cell = QTableWidgetItem(str(val))
                cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if col == 5:
                    colors = {"פעיל": COLORS['primary_light'], "הוחזר": COLORS['accent'], "באיחור": COLORS['danger']}
                    cell.setForeground(QColor(colors.get(val, COLORS['text_primary'])))
                self.history_table.setItem(row, col, cell)

    def _return_item(self, loan_id, item_id):
        reply = QMessageBox.question(self, "אישור החזרה", "לסמן חפץ זה כהוחזר?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.db.return_item(loan_id, item_id)
            self._refresh_active()
            self._refresh_overdue()
            self._refresh_history()
            QMessageBox.information(self, "בוצע", "החפץ סומן כהוחזר בהצלחה!")

    def _send_reminder(self, loan):
        dlg = EmailReminderDialog(
            self, loan['borrower_name'],
            loan.get('borrower_email', '') or '',
            loan['item_name'], loan['planned_return']
        )
        dlg.exec()

    def refresh(self):
        self._refresh_active()
        self._refresh_overdue()
        self._refresh_history()


class NewLoanPage(QWidget):
    loan_created = Signal()

    def __init__(self, db):
        super().__init__()
        self.db = db
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(20)

        title = QLabel("➕ השאלה חדשה")
        title.setObjectName("pageTitle")
        layout.addWidget(title)

        center = QHBoxLayout()
        center.addStretch()

        form_frame = card_frame()
        form_frame.setFixedWidth(520)
        form_layout = QVBoxLayout(form_frame)
        form_layout.setContentsMargins(28, 24, 28, 24)
        form_layout.setSpacing(18)

        subtitle = QLabel("מלא את פרטי ההשאלה")
        subtitle.setObjectName("sectionTitle")
        form_layout.addWidget(subtitle)
        form_layout.addWidget(sep_line())

        form = QFormLayout()
        form.setSpacing(14)
        form.setLabelAlignment(Qt.AlignRight)
        form.setFormAlignment(Qt.AlignRight | Qt.AlignTop)

        # Item
        self.item_search = QLineEdit()
        self.item_search.setPlaceholderText("🔍 חפש חפץ זמין...")
        self.item_search.textChanged.connect(self._filter_items)
        self.item_cb = QComboBox()
        self.item_cb.setMinimumHeight(36)
        self._load_items()
        form.addRow("חפץ *:", self.item_search)
        form.addRow("", self.item_cb)

        # Borrower
        self.bor_search = QLineEdit()
        self.bor_search.setPlaceholderText("🔍 חפש שואל...")
        self.bor_search.textChanged.connect(self._filter_borrowers)
        self.bor_cb = QComboBox()
        self.bor_cb.setMinimumHeight(36)
        self._load_borrowers()
        new_bor_btn = QPushButton("+ הוסף שואל חדש")
        new_bor_btn.setObjectName("secondaryBtn")
        new_bor_btn.clicked.connect(self._new_borrower)
        form.addRow("שואל *:", self.bor_search)
        form.addRow("", self.bor_cb)
        form.addRow("", new_bor_btn)

        # Dates
        self.return_date = QDateEdit()
        self.return_date.setCalendarPopup(True)
        self.return_date.setDate(QDate.currentDate().addDays(7))
        self.return_date.setMinimumDate(QDate.currentDate())
        self.return_date.setMinimumHeight(36)
        form.addRow("תאריך החזרה:", self.return_date)

        self.notes_edit = QLineEdit()
        self.notes_edit.setPlaceholderText("הערות אופציונליות...")
        form.addRow("הערות:", self.notes_edit)

        form_layout.addLayout(form)
        form_layout.addWidget(sep_line())

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        confirm_btn = QPushButton("✅  אשר השאלה")
        confirm_btn.setObjectName("successBtn")
        confirm_btn.setMinimumHeight(42)
        confirm_btn.setMinimumWidth(160)
        confirm_btn.clicked.connect(self._create_loan)
        btn_row.addWidget(confirm_btn)
        form_layout.addLayout(btn_row)

        center.addWidget(form_frame)
        center.addStretch()
        layout.addLayout(center)
        layout.addStretch()

    def _load_items(self, search=""):
        self.item_cb.clear()
        self._items = self.db.get_items(search=search, status="זמין")
        for it in self._items:
            self.item_cb.addItem(f"📦 {it['name']}  [{it['category']}]", it['id'])

    def _filter_items(self, text):
        self._load_items(search=text)

    def _load_borrowers(self, search=""):
        self.bor_cb.clear()
        self._borrowers = self.db.get_borrowers(search=search)
        for b in self._borrowers:
            stars = "⭐" * (b['rating'] or 0)
            self.bor_cb.addItem(f"👤 {b['full_name']}  |  {b['phone']}  {stars}", b['id'])

    def _filter_borrowers(self, text):
        self._load_borrowers(search=text)

    def _new_borrower(self):
        dlg = BorrowerDialog(self, self.db)
        if dlg.exec() == QDialog.Accepted:
            d = dlg.result_data
            bid = self.db.add_borrower(d['full_name'], d['phone'], d['email'], d['address'], d['rating'], d['notes'])
            self._load_borrowers()
            for i in range(self.bor_cb.count()):
                if self.bor_cb.itemData(i) == bid:
                    self.bor_cb.setCurrentIndex(i)
                    break

    def _create_loan(self):
        if self.item_cb.count() == 0:
            QMessageBox.warning(self, "שגיאה", "אין חפצים זמינים")
            return
        if self.bor_cb.count() == 0:
            QMessageBox.warning(self, "שגיאה", "יש לבחור שואל")
            return
        item_id = self.item_cb.currentData()
        borrower_id = self.bor_cb.currentData()
        planned_return = self.return_date.date().toString("yyyy-MM-dd")
        notes = self.notes_edit.text().strip()
        self.db.create_loan(item_id, borrower_id, planned_return, notes)
        self._load_items()
        QMessageBox.information(self, "בוצע!", "ההשאלה נוצרה בהצלחה! ✅")
        self.loan_created.emit()

    def refresh(self):
        self._load_items()
        self._load_borrowers()


class HebrewCalendarWidget(QWidget):
    """Hebrew/Gregorian dual-calendar widget powered by pyluach."""
    date_selected = Signal(object)

    GREG_MONTHS_HEB = ["ינואר","פברואר","מרץ","אפריל","מאי","יוני",
                       "יולי","אוגוסט","ספטמבר","אוקטובר","נובמבר","דצמבר"]
    DAY_HEADERS = ["ראשון","שני","שלישי","רביעי","חמישי","שישי","שבת"]

    def __init__(self, db, hebrew_mode=True):
        super().__init__()
        self.db = db
        self.hebrew_mode = hebrew_mode
        self._loan_dates = set()
        self._today = date.today()
        self._selected = self._today
        # current display: gregorian year+month
        self._cur_year  = self._today.year
        self._cur_month = self._today.month
        self._load_loan_dates()
        self._build_frame()
        self._render()

    def _load_loan_dates(self):
        rows = self.db.conn.execute(
            "SELECT planned_return FROM loans WHERE status IN ('פעיל','באיחור')"
        ).fetchall()
        self._loan_dates = {r[0] for r in rows}

    # ── skeleton (built once) ─────────────────────────────────────
    def _build_frame(self):
        from pyluach import dates as pdates
        main = QVBoxLayout(self)
        main.setContentsMargins(0, 0, 0, 0)
        main.setSpacing(0)

        nav = QHBoxLayout()
        nav.setContentsMargins(8, 8, 8, 4)
        self.prev_btn = QPushButton("◀")
        self.prev_btn.setFixedSize(32, 32)
        self.prev_btn.setObjectName("secondaryBtn")
        self.prev_btn.clicked.connect(self._prev_month)
        self.next_btn = QPushButton("▶")
        self.next_btn.setFixedSize(32, 32)
        self.next_btn.setObjectName("secondaryBtn")
        self.next_btn.clicked.connect(self._next_month)
        self.month_lbl = QLabel()
        self.month_lbl.setAlignment(Qt.AlignCenter)
        self.month_lbl.setStyleSheet("font-size:14px; font-weight:700;")
        nav.addWidget(self.next_btn)
        nav.addStretch()
        nav.addWidget(self.month_lbl)
        nav.addStretch()
        nav.addWidget(self.prev_btn)
        main.addLayout(nav)

        # Day-header row (7 fixed labels)
        hdr_row = QHBoxLayout()
        hdr_row.setContentsMargins(4, 0, 4, 2)
        hdr_row.setSpacing(2)
        for dh in self.DAY_HEADERS:
            lbl = QLabel(dh)
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setFixedHeight(24)
            lbl.setStyleSheet(
                f"font-size:10px; font-weight:600; color:{COLORS['text_secondary']};"
            )
            hdr_row.addWidget(lbl, 1)
        main.addLayout(hdr_row)

        # Grid — rebuilt every render
        self.grid_container = QWidget()
        self.grid_layout = QGridLayout(self.grid_container)
        self.grid_layout.setSpacing(2)
        self.grid_layout.setContentsMargins(4, 0, 4, 4)
        main.addWidget(self.grid_container)

    # ── render (called on month change / mode change) ─────────────
    def _render(self):
        from pyluach import dates as pdates
        import calendar as gcal

        # Clear grid
        while self.grid_layout.count():
            item = self.grid_layout.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

        y, m = self._cur_year, self._cur_month

        # Build header label
        greg_name = f"{self.GREG_MONTHS_HEB[m-1]} {y}"
        # Hebrew month of 1st of this greg month
        try:
            hd_first = pdates.HebrewDate.from_pydate(date(y, m, 1))
            heb_mon_name = hd_first.month_name(hebrew=True)
            heb_yr_str   = hd_first.hebrew_year()
            if self.hebrew_mode:
                self.month_lbl.setText(f"{heb_mon_name} {heb_yr_str}  |  {greg_name}")
            else:
                self.month_lbl.setText(greg_name)
        except Exception:
            self.month_lbl.setText(greg_name)

        # Iterate days
        days_in_month = gcal.monthrange(y, m)[1]
        row = 0
        for day in range(1, days_in_month + 1):
            g = date(y, m, day)
            date_str = g.isoformat()

            # column: isoweekday Mon=1..Sun=7 → Sun=col0..Sat=col6
            col = g.isoweekday() % 7  # Sun→0, Mon→1, ..., Sat→6

            is_today    = (g == self._today)
            is_selected = (g == self._selected)
            has_loan    = (date_str in self._loan_dates)

            # row index: when col wraps back to 0 after first week
            if day == 1:
                row = 0
            elif col == 0:
                row += 1

            # Hebrew date info
            try:
                hd = pdates.HebrewDate.from_pydate(g)
                heb_day_str = hd.hebrew_day()
                # add month name on 1st of hebrew month
                if hd.day == 1:
                    heb_day_str = f"{heb_day_str} {hd.month_name(hebrew=True)}"
            except Exception:
                heb_day_str = ""

            # Colors
            if is_selected:
                bg, fg, fg2 = COLORS['primary_light'], "white", "#cce6ff"
            elif is_today:
                bg, fg, fg2 = COLORS['badge_blue'], COLORS['primary'], COLORS['text_secondary']
            elif has_loan:
                bg, fg, fg2 = COLORS['badge_green'], COLORS['accent'], COLORS['text_secondary']
            else:
                bg, fg, fg2 = COLORS['surface'], COLORS['text_primary'], COLORS['text_secondary']

            cell = QFrame()
            cell.setFixedSize(66, 54)
            cell.setCursor(QCursor(Qt.PointingHandCursor))
            cell.setStyleSheet(
                f"background:{bg}; border-radius:7px; border:1px solid {COLORS['border']};"
            )
            cell.style().unpolish(cell)
            cell.style().polish(cell)
            cell.update()

            cl = QVBoxLayout(cell)
            cl.setContentsMargins(2, 2, 2, 2)
            cl.setSpacing(0)

            if self.hebrew_mode:
                center_txt = heb_day_str
                bottom_txt = str(day)
            else:
                center_txt = str(day)
                bottom_txt = heb_day_str

            lbl_center = QLabel(center_txt)
            lbl_center.setAlignment(Qt.AlignCenter)
            lbl_center.setStyleSheet(
                f"font-size:12px; font-weight:700; color:{fg}; background:transparent; border:none;"
            )
            lbl_bottom = QLabel(bottom_txt)
            lbl_bottom.setAlignment(Qt.AlignCenter)
            lbl_bottom.setStyleSheet(
                f"font-size:9px; color:{fg2}; background:transparent; border:none;"
            )

            if has_loan and not is_selected:
                dot = QLabel("●")
                dot.setAlignment(Qt.AlignCenter)
                dot.setStyleSheet(f"font-size:7px; color:{COLORS['accent']}; background:transparent; border:none;")
                cl.addWidget(dot)

            cl.addWidget(lbl_center)
            cl.addWidget(lbl_bottom)

            cell.mousePressEvent = lambda e, dt=g: self._cell_clicked(dt)
            self.grid_layout.addWidget(cell, row, col)

        self.grid_container.update()

    def _cell_clicked(self, dt):
        self._selected = dt
        self._render()
        self.date_selected.emit(dt)

    def _prev_month(self):
        if self._cur_month == 1:
            self._cur_month = 12
            self._cur_year -= 1
        else:
            self._cur_month -= 1
        self._render()

    def _next_month(self):
        if self._cur_month == 12:
            self._cur_month = 1
            self._cur_year += 1
        else:
            self._cur_month += 1
        self._render()

    def set_hebrew_mode(self, hebrew):
        self.hebrew_mode = hebrew
        self._render()


class CalendarPage(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        settings = db.get_settings()
        self._hebrew_mode = settings.get('calendar_mode', 'hebrew') == 'hebrew'
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        # Title row
        title_row = QHBoxLayout()
        title = QLabel("📅 לוח שנה - השאלות")
        title.setObjectName("pageTitle")
        title_row.addWidget(title)
        title_row.addStretch()
        layout.addLayout(title_row)

        splitter = QSplitter(Qt.Horizontal)

        # Calendar
        cal_frame = card_frame()
        cal_layout = QVBoxLayout(cal_frame)
        cal_layout.setContentsMargins(8, 8, 8, 8)
        self.cal_widget = HebrewCalendarWidget(self.db, hebrew_mode=self._hebrew_mode)
        self.cal_widget.date_selected.connect(self._date_selected)
        cal_layout.addWidget(self.cal_widget)
        splitter.addWidget(cal_frame)

        # Details panel
        detail_frame = card_frame()
        detail_layout = QVBoxLayout(detail_frame)
        detail_layout.setContentsMargins(16, 16, 16, 16)

        self.detail_title = QLabel("בחר תאריך לפרטים")
        self.detail_title.setObjectName("sectionTitle")
        detail_layout.addWidget(self.detail_title)
        detail_layout.addWidget(sep_line())

        self.detail_table = QTableWidget(0, 3)
        rtl_table(self.detail_table)
        self.detail_table.setHorizontalHeaderLabels(["חפץ", "שואל", "סטטוס"])
        self.detail_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.detail_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.detail_table.verticalHeader().setVisible(False)
        detail_layout.addWidget(self.detail_table)

        up_lbl = QLabel("📅  השאלות הקרובות (7 ימים)")
        up_lbl.setObjectName("sectionTitle")
        detail_layout.addWidget(up_lbl)

        self.upcoming_table = QTableWidget(0, 4)
        rtl_table(self.upcoming_table)
        self.upcoming_table.setHorizontalHeaderLabels(["חפץ", "שואל", "תאריך החזרה", "סטטוס"])
        self.upcoming_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.upcoming_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.upcoming_table.verticalHeader().setVisible(False)
        detail_layout.addWidget(self.upcoming_table)
        self._load_upcoming()

        splitter.addWidget(detail_frame)
        splitter.setSizes([560, 380])
        layout.addWidget(splitter)

    def _date_selected(self, selected_date):
        date_str = selected_date.isoformat()
        self.detail_title.setText(f"השאלות לתאריך: {selected_date.strftime('%d/%m/%Y')}")

        loans = self.db.conn.execute("""
            SELECT l.*, i.name as item_name, b.full_name as borrower_name
            FROM loans l JOIN items i ON l.item_id=i.id JOIN borrowers b ON l.borrower_id=b.id
            WHERE l.planned_return=? OR date(l.loan_date)=?
        """, (date_str, date_str)).fetchall()

        self.detail_table.setRowCount(len(loans))
        for row, loan in enumerate(loans):
            for col, val in enumerate([loan['item_name'], loan['borrower_name'], loan['status']]):
                cell = QTableWidgetItem(val)
                cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.detail_table.setItem(row, col, cell)

    def _load_upcoming(self):
        today = date.today()
        in7 = today + timedelta(days=7)
        loans = self.db.conn.execute("""
            SELECT l.*, i.name as item_name, b.full_name as borrower_name
            FROM loans l JOIN items i ON l.item_id=i.id JOIN borrowers b ON l.borrower_id=b.id
            WHERE l.planned_return BETWEEN ? AND ? AND l.status IN ('פעיל','באיחור')
            ORDER BY l.planned_return
        """, (today.isoformat(), in7.isoformat())).fetchall()

        self.upcoming_table.setRowCount(len(loans))
        for row, loan in enumerate(loans):
            for col, val in enumerate([loan['item_name'], loan['borrower_name'], loan['planned_return'], loan['status']]):
                cell = QTableWidgetItem(str(val))
                cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.upcoming_table.setItem(row, col, cell)

    def set_hebrew_mode(self, hebrew):
        self._hebrew_mode = hebrew
        self.cal_widget.set_hebrew_mode(hebrew)

    def refresh(self):
        self._load_upcoming()
        self.cal_widget._load_loan_dates()
        self.cal_widget._render()
        self.cal_widget.update()
        self.update()


class ReportsBridge(QObject):
    """JS → Python bridge for the Reports page export buttons."""
    exportRequested = Signal(str)

    @Slot(str)
    def export(self, kind):
        self.exportRequested.emit(kind)


class ReportsPage(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self._use_web = WEBENGINE_AVAILABLE
        self._web_view = None
        self._bridge = None

    def refresh(self):
        if self._use_web:
            self._render_web()
        else:
            self._build_classic()

    # ── HTML/CSS rich reports ───────────────────────────────────────
    def _ensure_web_view(self):
        if self._web_view is not None:
            return
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        self._web_view = QWebEngineView()
        self._web_view.setContextMenuPolicy(Qt.NoContextMenu)
        layout.addWidget(self._web_view)

        self._bridge = ReportsBridge()
        self._bridge.exportRequested.connect(self._handle_export)
        self._channel = QWebChannel()
        self._channel.registerObject("bridge", self._bridge)
        self._web_view.page().setWebChannel(self._channel)

    def _handle_export(self, kind):
        if kind == "items":
            self._export_items_excel()
        elif kind == "loans":
            self._export_loans_excel()
        elif kind == "borrowers":
            self._export_borrowers_excel()

    def _render_web(self):
        try:
            self._ensure_web_view()
            html = self._build_html()
            self._web_view.setHtml(html, baseUrl=QUrl("qrc:/"))
        except Exception:
            self._use_web = False
            self._build_classic()

    def _build_html(self):
        stats = self.db.get_stats()

        cat_rows = ""
        max_total = 1
        cat_data = []
        for cat in CATEGORIES:
            total = self.db.conn.execute("SELECT COUNT(*) FROM items WHERE category=?", (cat,)).fetchone()[0]
            avail = self.db.conn.execute("SELECT COUNT(*) FROM items WHERE category=? AND status='זמין'", (cat,)).fetchone()[0]
            loaned = self.db.conn.execute("SELECT COUNT(*) FROM items WHERE category=? AND status='מושאל'", (cat,)).fetchone()[0]
            cat_data.append((cat, total, avail, loaned))
            max_total = max(max_total, total)

        for cat, total, avail, loaned in cat_data:
            bar_pct = round(total / max_total * 100) if max_total else 0
            avail_pct = round(avail / total * 100) if total else 0
            cat_rows += f"""
            <tr>
              <td class="cat-name">{DashboardPage._esc(cat)}</td>
              <td class="num">{total}</td>
              <td class="num accent-green">{avail}</td>
              <td class="num accent-navy">{loaned}</td>
              <td class="bar-cell">
                <div class="bar-track">
                  <div class="bar-fill" style="width:{bar_pct}%"></div>
                </div>
              </td>
            </tr>"""

        html = f"""<!DOCTYPE html>
<html lang="he" dir="rtl">
<head>
<meta charset="UTF-8">
<style>
{self._css()}
</style>
</head>
<body>
<div class="wrap">

  <header class="page-head">
    <h1>📊 דוחות וסטטיסטיקות</h1>
    <p class="subtitle">תמונת מצב כללית של הגמ"ח וייצוא נתונים לקובצי Excel</p>
  </header>

  <section class="stat-band">
    <div class="stat-block">
      <span class="stat-val accent-navy">{stats['total_loans']}</span>
      <span class="stat-label">סה"כ השאלות</span>
    </div>
    <div class="stat-sep"></div>
    <div class="stat-block">
      <span class="stat-val accent-green">{stats['available']}</span>
      <span class="stat-label">חפצים זמינים</span>
    </div>
    <div class="stat-sep"></div>
    <div class="stat-block">
      <span class="stat-val accent-gold">{stats['loaned']}</span>
      <span class="stat-label">מושאלים</span>
    </div>
    <div class="stat-sep"></div>
    <div class="stat-block">
      <span class="stat-val accent-red">{stats['overdue']}</span>
      <span class="stat-label">באיחור</span>
    </div>
  </section>

  <section class="panel">
    <div class="panel-head"><h2>📤 ייצוא נתונים ל-Excel</h2></div>
    <div class="export-row">
      <button class="export-btn" onclick="doExport('items')">
        <span class="export-icon">📗</span>
        <span class="export-text">ייצוא חפצים</span>
      </button>
      <button class="export-btn" onclick="doExport('loans')">
        <span class="export-icon">📗</span>
        <span class="export-text">ייצוא השאלות</span>
      </button>
      <button class="export-btn" onclick="doExport('borrowers')">
        <span class="export-icon">📗</span>
        <span class="export-text">ייצוא שואלים</span>
      </button>
    </div>
  </section>

  <section class="panel">
    <div class="panel-head"><h2>📦 פירוט לפי קטגוריה</h2></div>
    <table class="ledger">
      <thead>
        <tr><th>קטגוריה</th><th>סה"כ חפצים</th><th>זמינים</th><th>מושאלים</th><th>יחס מהמלאי</th></tr>
      </thead>
      <tbody>{cat_rows}</tbody>
    </table>
  </section>

</div>

<script src="qrc:///qtwebchannel/qwebchannel.js"></script>
<script>
  let bridge = null;
  new QWebChannel(qt.webChannelTransport, function(channel) {{
    bridge = channel.objects.bridge;
  }});
  function doExport(kind) {{
    if (bridge) {{ bridge.export(kind); }}
  }}
</script>
</body>
</html>"""
        return html

    @staticmethod
    def _css():
        root_vars = f"""
        :root {{
          --navy: {COLORS.get('html_navy', '#0F1B2E')};
          --navy-2: {COLORS.get('html_navy2', '#16263F')};
          --gold: {COLORS.get('html_gold', '#D4A24C')};
          --gold-light: {COLORS.get('html_gold_light', '#E8C588')};
          --sage: {COLORS.get('html_sage', '#5B8C6E')};
          --sage-light: {COLORS.get('html_sage_light', '#E3EEE6')};
          --rose: {COLORS.get('html_rose', '#C75D4D')};
          --rose-light: {COLORS.get('html_rose_light', '#F7E2DE')};
          --parchment: {COLORS.get('html_parchment', '#F7F3EC')};
          --paper: {COLORS.get('card', '#FFFFFF')};
          --ink: {COLORS.get('text_primary', '#1F2A33')};
          --ink-soft: {COLORS.get('text_secondary', '#6B7780')};
          --line: {COLORS.get('html_line', '#E7E0D2')};
        }}
        """
        rest = """
        * { box-sizing: border-box; }
        body {
          margin: 0;
          background: var(--parchment);
          font-family: "Segoe UI", Tahoma, Arial, sans-serif;
          color: var(--ink);
        }
        .wrap { padding: 26px 30px 40px; max-width: 1180px; margin: 0 auto; }

        .page-head { margin-bottom: 18px; animation: fadeUp .5s ease both; }
        .page-head h1 { font-size: 24px; margin: 0 0 4px; font-weight: 700; }
        .subtitle { color: var(--ink-soft); font-size: 13px; margin: 0; }

        .stat-band {
          background: linear-gradient(135deg, var(--navy) 0%, var(--navy-2) 60%, #1d3450 100%);
          border-radius: 16px;
          padding: 22px 10px;
          display: flex;
          align-items: center;
          justify-content: space-around;
          box-shadow: 0 10px 26px rgba(15,27,46,0.22);
          margin-bottom: 18px;
          animation: fadeUp .55s ease both;
          animation-delay: .05s;
        }
        .stat-block { display: flex; flex-direction: column; align-items: center; gap: 4px; flex: 1; }
        .stat-val { font-size: 27px; font-weight: 700; color: #fff; }
        .stat-label { font-size: 12px; color: #B9C4D4; }
        .stat-sep { width: 1px; height: 38px; background: rgba(255,255,255,0.18); }
        .accent-green { color: var(--sage) !important; }
        .accent-navy { color: var(--gold-light) !important; }
        .accent-gold { color: var(--gold-light) !important; }
        .accent-red { color: #E69A8D !important; }

        .panel {
          background: var(--paper);
          border: 1px solid var(--line);
          border-radius: 16px;
          margin-bottom: 18px;
          padding: 18px 20px;
          box-shadow: 0 2px 10px rgba(15,27,46,0.04);
          animation: fadeUp .6s ease both;
          animation-delay: .1s;
        }
        .panel-head { margin-bottom: 14px; }
        .panel-head h2 { font-size: 15.5px; margin: 0; font-weight: 700; }

        .export-row { display: flex; gap: 12px; flex-wrap: wrap; }
        .export-btn {
          flex: 1;
          min-width: 160px;
          background: linear-gradient(135deg, var(--navy), var(--navy-2));
          color: #fff;
          border: none;
          border-radius: 12px;
          padding: 16px 14px;
          display: flex;
          align-items: center;
          gap: 10px;
          cursor: pointer;
          font-family: inherit;
          font-size: 13.5px;
          font-weight: 600;
          transition: transform .15s ease, box-shadow .15s ease;
        }
        .export-btn:hover {
          transform: translateY(-2px);
          box-shadow: 0 8px 18px rgba(15,27,46,0.25);
        }
        .export-icon { font-size: 20px; }

        .ledger { width: 100%; border-collapse: collapse; font-size: 13px; }
        .ledger th {
          text-align: right; color: var(--ink-soft); font-weight: 600;
          font-size: 11.5px; padding: 8px; border-bottom: 1px solid var(--line);
        }
        .ledger td { padding: 11px 8px; border-bottom: 1px solid var(--line); vertical-align: middle; }
        .ledger tr:hover td { background: var(--parchment); }
        .cat-name { font-weight: 600; }
        .num { color: var(--ink); font-weight: 600; }
        .num.accent-green { color: var(--sage); }
        .num.accent-navy { color: var(--navy); }
        .bar-cell { width: 30%; }
        .bar-track { background: var(--line); border-radius: 999px; height: 8px; overflow: hidden; }
        .bar-fill { background: linear-gradient(90deg, var(--gold), var(--sage)); height: 100%; border-radius: 999px; transition: width 1s ease; }

        @keyframes fadeUp {
          from { opacity: 0; transform: translateY(10px); }
          to { opacity: 1; transform: translateY(0); }
        }
        """
        return root_vars + rest

    # ── classic Qt-widgets fallback ─────────────────────────────────
    def _build_classic(self):
        old_layout = self.layout()
        if old_layout is not None:
            while old_layout.count():
                item = old_layout.takeAt(0)
                w = item.widget()
                if w:
                    w.setParent(None)
                    w.deleteLater()
            import shiboken6
            shiboken6.delete(old_layout)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        title = QLabel("📊 דוחות וסטטיסטיקות")
        title.setObjectName("pageTitle")
        layout.addWidget(title)

        stats = self.db.get_stats()
        stats_frame = card_frame()
        stats_fl = QHBoxLayout(stats_frame)
        stats_fl.setContentsMargins(20, 16, 20, 16)

        for label, val, color in [
            ("סה״כ השאלות", stats['total_loans'], COLORS['primary_light']),
            ("חפצים זמינים", stats['available'], COLORS['accent']),
            ("מושאלים", stats['loaned'], COLORS['warning']),
            ("באיחור", stats['overdue'], COLORS['danger']),
        ]:
            v_layout = QVBoxLayout()
            v_lbl = QLabel(str(val))
            v_lbl.setStyleSheet(f"font-size: 26px; font-weight: 700; color: {color};")
            v_lbl.setAlignment(Qt.AlignCenter)
            t_lbl = QLabel(label)
            t_lbl.setObjectName("statLabel")
            t_lbl.setAlignment(Qt.AlignCenter)
            v_layout.addWidget(v_lbl)
            v_layout.addWidget(t_lbl)
            stats_fl.addLayout(v_layout)
            if label != "באיחור":
                sep = QFrame()
                sep.setFrameShape(QFrame.VLine)
                sep.setStyleSheet(f"color: {COLORS['border']};")
                stats_fl.addWidget(sep)

        layout.addWidget(stats_frame)

        export_frame = card_frame()
        export_layout = QVBoxLayout(export_frame)
        export_layout.setContentsMargins(20, 16, 20, 16)

        exp_title = QLabel("📤 ייצוא נתונים")
        exp_title.setObjectName("sectionTitle")
        export_layout.addWidget(exp_title)

        btn_row = QHBoxLayout()
        for text, fn in [
            ("📗 ייצוא חפצים ל-Excel", self._export_items_excel),
            ("📗 ייצוא השאלות ל-Excel", self._export_loans_excel),
            ("📗 ייצוא שואלים ל-Excel", self._export_borrowers_excel),
        ]:
            btn = QPushButton(text)
            btn.setObjectName("primaryBtn")
            btn.clicked.connect(fn)
            btn_row.addWidget(btn)
        btn_row.addStretch()
        export_layout.addLayout(btn_row)
        layout.addWidget(export_frame)

        cat_frame = card_frame()
        cat_layout = QVBoxLayout(cat_frame)
        cat_layout.setContentsMargins(20, 16, 20, 16)
        cat_title = QLabel("📦 פירוט לפי קטגוריה")
        cat_title.setObjectName("sectionTitle")
        cat_layout.addWidget(cat_title)

        cat_tbl = QTableWidget(0, 4)
        rtl_table(cat_tbl)
        cat_tbl.setHorizontalHeaderLabels(["קטגוריה", "סה״כ חפצים", "זמינים", "מושאלים"])
        cat_tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        cat_tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        cat_tbl.verticalHeader().setVisible(False)

        cat_tbl.setRowCount(len(CATEGORIES))
        for row, cat in enumerate(CATEGORIES):
            total = self.db.conn.execute("SELECT COUNT(*) FROM items WHERE category=?", (cat,)).fetchone()[0]
            avail = self.db.conn.execute("SELECT COUNT(*) FROM items WHERE category=? AND status='זמין'", (cat,)).fetchone()[0]
            loaned = self.db.conn.execute("SELECT COUNT(*) FROM items WHERE category=? AND status='מושאל'", (cat,)).fetchone()[0]
            for col, val in enumerate([cat, str(total), str(avail), str(loaned)]):
                cell = rtl_item(val)
                cat_tbl.setItem(row, col, cell)

        cat_tbl.setFixedHeight(200)
        cat_layout.addWidget(cat_tbl)
        layout.addWidget(cat_frame)
        layout.addStretch()

    def _export_items_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "שמור קובץ Excel", "חפצים.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "חפצים"
            ws.sheet_view.rightToLeft = True

            headers = ["מזהה", "שם החפץ", "קטגוריה", "סטטוס", "מיקום", "מספר סידורי", "הערות"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(1, col, h)
                cell.font = Font(bold=True, name="Arial")
                cell.fill = PatternFill("solid", fgColor="1B4F72")
                cell.font = Font(bold=True, color="FFFFFF", name="Arial")
                cell.alignment = Alignment(horizontal="right")

            for row, it in enumerate(self.db.get_items(), 2):
                for col, val in enumerate([
                    it['id'], it['name'], it['category'], it['status'],
                    it['location'] or "", it['serial_number'] or "", it['notes'] or ""
                ], 1):
                    c = ws.cell(row, col, val)
                    c.alignment = Alignment(horizontal="right")
                    c.font = Font(name="Arial")

            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col) + 4
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 40)

            wb.save(path)
            QMessageBox.information(self, "הצלחה", f"הקובץ נשמר:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "שגיאה", str(e))

    def _export_loans_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "שמור", "השאלות.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "השאלות"
            ws.sheet_view.rightToLeft = True

            headers = ["מזהה", "חפץ", "שואל", "טלפון", "תאריך השאלה", "תאריך החזרה", "הוחזר", "סטטוס"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(1, col, h)
                cell.font = Font(bold=True, color="FFFFFF", name="Arial")
                cell.fill = PatternFill("solid", fgColor="27AE60")
                cell.alignment = Alignment(horizontal="right")

            for row, loan in enumerate(self.db.get_loans(), 2):
                for col, val in enumerate([
                    loan['id'], loan['item_name'], loan['borrower_name'],
                    loan['phone'], loan['loan_date'][:10] if loan['loan_date'] else "",
                    loan['planned_return'],
                    loan['actual_return'][:10] if loan['actual_return'] else "",
                    loan['status']
                ], 1):
                    c = ws.cell(row, col, val)
                    c.alignment = Alignment(horizontal="right")
                    c.font = Font(name="Arial")

            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col) + 4, 40
                )
            wb.save(path)
            QMessageBox.information(self, "הצלחה", f"הקובץ נשמר:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "שגיאה", str(e))

    def _export_borrowers_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "שמור", "שואלים.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "שואלים"
            ws.sheet_view.rightToLeft = True

            headers = ["מזהה", "שם מלא", "טלפון", "דוא״ל", "כתובת", "דירוג"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(1, col, h)
                cell.font = Font(bold=True, color="FFFFFF", name="Arial")
                cell.fill = PatternFill("solid", fgColor="2E86C1")
                cell.alignment = Alignment(horizontal="right")

            for row, b in enumerate(self.db.get_borrowers(), 2):
                for col, val in enumerate([
                    b['id'], b['full_name'], b['phone'],
                    b['email'] or "", b['address'] or "", b['rating'] or 0
                ], 1):
                    c = ws.cell(row, col, val)
                    c.alignment = Alignment(horizontal="right")
                    c.font = Font(name="Arial")

            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col) + 4, 40
                )
            wb.save(path)
            QMessageBox.information(self, "הצלחה", f"הקובץ נשמר:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "שגיאה", str(e))

class EmailHelpDialog(QDialog):
    """Guided explanation of how to set up email reminders (Gmail App Password etc.)."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("מדריך הגדרת מייל לתזכורות")
        self.setLayoutDirection(Qt.RightToLeft)
        self.setMinimumWidth(560)
        self.setMinimumHeight(520)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(14)

        title = QLabel("📧 איך מגדירים שליחת תזכורות במייל?")
        title.setObjectName("pageTitle")
        layout.addWidget(title)
        layout.addWidget(sep_line())

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        content = QWidget()
        cl = QVBoxLayout(content)
        cl.setSpacing(14)

        steps = [
            ("1️⃣ למה צריך \"App Password\" ולא את הסיסמה הרגילה?",
             "חברות מייל כמו Gmail חוסמות התחברות תוכנות חיצוניות עם הסיסמה "
             "הרגילה מטעמי אבטחה. יש ליצור \"סיסמת אפליקציה\" (App Password) "
             "מיוחדת המשמשת רק את התוכנה הזו, ואינה חושפת את הסיסמה האמיתית "
             "שלך לחשבון."),
            ("2️⃣ הפעלת אימות דו-שלבי (2-Step Verification)",
             "בחשבון Google שלך: היכנסו ל-myaccount.google.com ← אבטחה ← "
             "אימות בשני שלבים. יש להפעיל אפשרות זו לפני שניתן ליצור סיסמת "
             "אפליקציה."),
            ("3️⃣ יצירת סיסמת אפליקציה",
             "לאחר הפעלת האימות הדו-שלבי, חפשו בדף האבטחה את \"סיסמאות "
             "אפליקציה\" (App Passwords). בחרו שם לאפליקציה (לדוגמה: גמ\"ח), "
             "ולחצו על יצירה. Google תציג קוד בן 16 תווים — זהו הקוד שיש "
             "להדביק בשדה \"סיסמה\" בתוכנה (לא הסיסמה הרגילה לחשבון!)."),
            ("4️⃣ מה למלא בשדות התוכנה?",
             "שרת SMTP: smtp.gmail.com   |   פורט: 587\n"
             "מייל: כתובת ה-Gmail המלאה שלך (לדוגמה: gemach@gmail.com)\n"
             "סיסמה: 16 התווים שהתקבלו בשלב הקודם, בלי רווחים."),
            ("5️⃣ ספקי מייל אחרים",
             "Outlook / Hotmail: smtp.office365.com, פורט 587\n"
             "Yahoo: smtp.mail.yahoo.com, פורט 587 (גם כאן נדרשת סיסמת אפליקציה)\n"
             "אם אינכם בטוחים, חפשו \"שם הספק + SMTP settings\"."),
            ("6️⃣ בדיקה",
             "לאחר מילוי הפרטים, השתמשו בכפתור \"בדיקת שליחה\" בעמוד התזכורות "
             "כדי לאמת שההגדרות תקינות לפני שתסמכו על השליחה האוטומטית."),
        ]

        for heading, body in steps:
            box = card_frame()
            bl = QVBoxLayout(box)
            bl.setContentsMargins(16, 12, 16, 14)
            h = QLabel(heading)
            h.setObjectName("sectionTitle")
            h.setWordWrap(True)
            bl.addWidget(h)
            b = QLabel(body)
            b.setWordWrap(True)
            b.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 12.5px; line-height: 1.5;")
            bl.addWidget(b)
            cl.addWidget(box)

        cl.addStretch()
        scroll.setWidget(content)
        layout.addWidget(scroll)

        close_btn = QPushButton("הבנתי, סגור")
        close_btn.setObjectName("primaryBtn")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)


class RemindersPage(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self._build()

    def _build(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        outer.addWidget(scroll)

        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)
        scroll.setWidget(content)

        title = QLabel("🔔 תזכורות")
        title.setObjectName("pageTitle")
        layout.addWidget(title)

        settings = self.db.get_settings()

        # ── Email / SMTP settings ──
        email_frame = card_frame()
        el = QVBoxLayout(email_frame)
        el.setContentsMargins(20, 16, 20, 20)
        el.setSpacing(10)

        head_row = QHBoxLayout()
        head_lbl = QLabel("📧 הגדרות מייל אוטומטי")
        head_lbl.setObjectName("sectionTitle")
        head_row.addWidget(head_lbl)
        head_row.addStretch()
        help_btn = QPushButton("❓ מדריך הגדרה")
        help_btn.setObjectName("secondaryBtn")
        help_btn.clicked.connect(self._show_help)
        head_row.addWidget(help_btn)
        el.addLayout(head_row)
        el.addWidget(sep_line())

        form = QFormLayout()
        form.setSpacing(12)
        form.setLabelAlignment(Qt.AlignRight)
        form.setFormAlignment(Qt.AlignRight | Qt.AlignTop)

        self.smtp_server = QLineEdit(settings.get('smtp_server', 'smtp.gmail.com'))
        self.smtp_port = QLineEdit(settings.get('smtp_port', '587'))
        self.smtp_user = QLineEdit(settings.get('smtp_user', ''))
        self.smtp_user.setPlaceholderText("your@gmail.com")
        self.smtp_pass = QLineEdit(settings.get('smtp_pass', ''))
        self.smtp_pass.setEchoMode(QLineEdit.Password)
        self.smtp_pass.setPlaceholderText("App Password (16 תווים)")

        form.addRow("שרת SMTP:", self.smtp_server)
        form.addRow("פורט:", self.smtp_port)
        form.addRow("מייל שולח:", self.smtp_user)
        form.addRow("סיסמה / App Password:", self.smtp_pass)

        self.reminder_days = QSpinBox()
        self.reminder_days.setRange(1, 14)
        self.reminder_days.setValue(int(settings.get('reminder_days', 2)))
        self.reminder_days.setSuffix(" ימים לפני המועד")
        form.addRow("שליחת תזכורת אוטומטית:", self.reminder_days)

        el.addLayout(form)

        btn_row = QHBoxLayout()
        save_email_btn = QPushButton("💾 שמור הגדרות")
        save_email_btn.setObjectName("primaryBtn")
        save_email_btn.clicked.connect(self._save_email)
        test_btn = QPushButton("✉️ בדיקת שליחה")
        test_btn.setObjectName("secondaryBtn")
        test_btn.clicked.connect(self._test_send)
        btn_row.addWidget(save_email_btn)
        btn_row.addWidget(test_btn)
        btn_row.addStretch()
        el.addLayout(btn_row)

        layout.addWidget(email_frame)

        # ── Manual reminders list ──
        list_frame = card_frame()
        ll = QVBoxLayout(list_frame)
        ll.setContentsMargins(20, 16, 20, 16)
        ll_title = QLabel("📋 שליחת תזכורת ידנית לשואל")
        ll_title.setObjectName("sectionTitle")
        ll.addWidget(ll_title)
        ll_sub = QLabel("רשימת ההשאלות הפעילות והבאות לאיחור — ניתן לשלוח תזכורת ידנית לכל שואל")
        ll_sub.setStyleSheet(f"color:{COLORS['text_secondary']}; font-size:12px;")
        ll_sub.setWordWrap(True)
        ll.addWidget(ll_sub)

        self.reminders_table = QTableWidget(0, 6)
        rtl_table(self.reminders_table)
        self.reminders_table.setHorizontalHeaderLabels(
            ["חפץ", "שואל", "טלפון", "מייל", "להחזרה עד", "פעולות"]
        )
        self.reminders_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.reminders_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.reminders_table.horizontalHeader().setMinimumSectionSize(100)
        self.reminders_table.setColumnWidth(2, 110)
        self.reminders_table.setColumnWidth(3, 150)
        self.reminders_table.setColumnWidth(4, 110)
        self.reminders_table.setColumnWidth(5, 150)
        self.reminders_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.reminders_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.reminders_table.verticalHeader().setVisible(False)
        self.reminders_table.setMinimumHeight(260)
        ll.addWidget(self.reminders_table)
        layout.addWidget(list_frame)

        layout.addStretch()
        self._refresh_table()

    def _refresh_table(self):
        loans = self.db.get_loans(active_only=True)
        self.reminders_table.setRowCount(len(loans))
        today = date.today()
        for row, loan in enumerate(loans):
            self.reminders_table.setRowHeight(row, 46)
            is_overdue = loan['status'] == 'באיחור'
            for col, val in enumerate([
                loan['item_name'], loan['borrower_name'], loan['phone'],
                loan['borrower_email'] or "—", loan['planned_return']
            ]):
                color = COLORS['danger'] if (is_overdue and col == 4) else None
                cell = rtl_item(val, color)
                self.reminders_table.setItem(row, col, cell)

            aw = QWidget()
            aw.setLayoutDirection(Qt.RightToLeft)
            al = QHBoxLayout(aw)
            al.setContentsMargins(4, 2, 4, 2)
            al.setSpacing(4)
            send_btn = QPushButton("📧 שלח תזכורת")
            send_btn.setObjectName("primaryBtn" if not is_overdue else "dangerBtn")
            send_btn.setFixedHeight(30)
            send_btn.setMinimumWidth(120)
            send_btn.clicked.connect(lambda _, l=loan: self._send_reminder(l))
            al.addWidget(send_btn)
            self.reminders_table.setCellWidget(row, 5, aw)

    def _send_reminder(self, loan):
        dlg = EmailReminderDialog(
            self, loan['borrower_name'],
            loan.get('borrower_email', '') or '',
            loan['item_name'], loan['planned_return']
        )
        # Pre-fill sender credentials from saved settings for convenience
        settings = self.db.get_settings()
        dlg.smtp_server.setText(settings.get('smtp_server', 'smtp.gmail.com'))
        dlg.smtp_port.setText(settings.get('smtp_port', '587'))
        dlg.sender_email.setText(settings.get('smtp_user', ''))
        dlg.sender_pass.setText(settings.get('smtp_pass', ''))
        dlg.exec()

    def _save_email(self):
        self.db.save_setting('smtp_server', self.smtp_server.text())
        self.db.save_setting('smtp_port', self.smtp_port.text())
        self.db.save_setting('smtp_user', self.smtp_user.text())
        self.db.save_setting('smtp_pass', self.smtp_pass.text())
        self.db.save_setting('reminder_days', str(self.reminder_days.value()))
        QMessageBox.information(self, "הצלחה", "ההגדרות נשמרו!")

    def _test_send(self):
        server = self.smtp_server.text().strip()
        port = self.smtp_port.text().strip()
        user = self.smtp_user.text().strip()
        pwd = self.smtp_pass.text().strip()

        if not (server and port and user and pwd):
            QMessageBox.warning(self, "שגיאה", "יש למלא את כל שדות המייל לפני הבדיקה")
            return

        try:
            port_int = int(port)
        except ValueError:
            QMessageBox.warning(self, "שגיאה", "הפורט חייב להיות מספר (לדוגמה 587)")
            return

        try:
            with smtplib.SMTP(server, port_int, timeout=10) as s:
                s.starttls()
                s.login(user, pwd)
            QMessageBox.information(self, "הצלחה",
                "החיבור לשרת המייל הצליח! ניתן לשלוח תזכורות מהגדרות אלו.")
        except Exception as e:
            QMessageBox.critical(self, "שגיאה בחיבור",
                f"לא ניתן להתחבר עם הפרטים שסופקו:\n{e}\n\n"
                "בדקו שם משתמש, סיסמה (App Password), שרת ופורט.\n"
                "לחצו על 'מדריך הגדרה' להוראות מפורטות.")

    def _show_help(self):
        dlg = EmailHelpDialog(self)
        dlg.exec()

    def refresh(self):
        self._refresh_table()



class SettingsPage(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self._theme_buttons = {}
        self._build()

    def _build(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        outer.addWidget(scroll)

        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)
        scroll.setWidget(content)

        title = QLabel("⚙️ הגדרות")
        title.setObjectName("pageTitle")
        layout.addWidget(title)

        settings = self.db.get_settings()

        # ── Appearance / theme ──
        appearance_frame = card_frame()
        af = QVBoxLayout(appearance_frame)
        af.setContentsMargins(20, 16, 20, 20)
        af.setSpacing(12)
        appearance_title = QLabel("🎨 עיצוב ומראה התוכנה")
        appearance_title.setObjectName("sectionTitle")
        af.addWidget(appearance_title)
        af.addWidget(sep_line())

        sub = QLabel("בחר/י סכמת צבעים לכל התוכנה — לוח הבקרה, הדוחות, הטבלאות והכפתורים:")
        sub.setWordWrap(True)
        sub.setStyleSheet(f"color:{COLORS['text_secondary']}; font-size:12.5px;")
        af.addWidget(sub)

        swatch_row = QHBoxLayout()
        swatch_row.setSpacing(12)
        current_theme = settings.get('theme', DEFAULT_THEME)
        for theme_key, theme_data in THEMES.items():
            btn = self._make_theme_swatch(theme_key, theme_data, current_theme == theme_key)
            self._theme_buttons[theme_key] = btn
            swatch_row.addWidget(btn)
        swatch_row.addStretch()
        af.addLayout(swatch_row)

        layout.addWidget(appearance_frame)

        # ── Calendar mode ──
        cal_frame2 = card_frame()
        cal_l = QVBoxLayout(cal_frame2)
        cal_l.setContentsMargins(20, 16, 20, 16)
        cal_title2 = QLabel("📅 הגדרות לוח שנה")
        cal_title2.setObjectName("sectionTitle")
        cal_l.addWidget(cal_title2)
        cal_l.addWidget(sep_line())
        cal_mode_row = QHBoxLayout()
        cal_mode_lbl = QLabel("פריסת לוח שנה:")
        self.cal_mode_cb = QComboBox()
        self.cal_mode_cb.addItem("לוח עברי (ברירת מחדל)", "hebrew")
        self.cal_mode_cb.addItem("לוח לועזי", "gregorian")
        saved_mode = settings.get('calendar_mode', 'hebrew')
        self.cal_mode_cb.setCurrentIndex(0 if saved_mode == 'hebrew' else 1)
        cal_mode_row.addWidget(cal_mode_lbl)
        cal_mode_row.addWidget(self.cal_mode_cb)
        cal_mode_row.addStretch()
        cal_l.addLayout(cal_mode_row)
        save_cal_btn = QPushButton("💾 שמור הגדרת לוח שנה")
        save_cal_btn.setObjectName("primaryBtn")
        save_cal_btn.clicked.connect(self._save_calendar_mode)
        cal_l.addWidget(save_cal_btn)
        layout.addWidget(cal_frame2)

        # ── DB info ──
        db_frame = card_frame()
        dl = QVBoxLayout(db_frame)
        dl.setContentsMargins(20, 16, 20, 16)
        db_title = QLabel("🗄️ מידע על מסד הנתונים")
        db_title.setObjectName("sectionTitle")
        dl.addWidget(db_title)
        dl.addWidget(sep_line())
        dl.addWidget(QLabel(f"מיקום קובץ: {DB_PATH}"))
        dl.addWidget(QLabel(f"תיקיית תמונות: {IMAGES_DIR}"))

        backup_btn = QPushButton("💾 גיבוי מסד הנתונים")
        backup_btn.setObjectName("secondaryBtn")
        backup_btn.clicked.connect(self._backup_db)
        dl.addWidget(backup_btn)
        layout.addWidget(db_frame)

        layout.addStretch()

    def _make_theme_swatch(self, theme_key, theme_data, is_active):
        btn = QPushButton()
        btn.setCheckable(True)
        btn.setChecked(is_active)
        btn.setFixedSize(140, 92)
        btn.setCursor(QCursor(Qt.PointingHandCursor))
        btn.setProperty("themeKey", theme_key)

        border = f"3px solid {theme_data['primary_light']}" if is_active else f"1.5px solid {COLORS['border']}"
        btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
                    stop:0 {theme_data['sidebar']}, stop:0.55 {theme_data['primary_light']},
                    stop:1 {theme_data.get('accent', theme_data['primary_light'])});
                border: {border};
                border-radius: 12px;
                color: white;
                font-size: 11px;
                font-weight: 600;
                padding: 6px;
                text-align: center;
            }}
            QPushButton:hover {{
                border: 3px solid {theme_data['primary_light']};
            }}
        """)
        btn.setText(("✓  " if is_active else "") + theme_data['display_name'])
        btn.clicked.connect(lambda: self._select_theme(theme_key))
        return btn

    def _select_theme(self, theme_key):
        self.db.save_setting('theme', theme_key)
        for key, btn in self._theme_buttons.items():
            is_active = (key == theme_key)
            btn.setChecked(is_active)
            theme_data = THEMES[key]
            border = f"3px solid {theme_data['primary_light']}" if is_active else "1.5px solid #E2E8F0"
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
                        stop:0 {theme_data['sidebar']}, stop:0.55 {theme_data['primary_light']},
                        stop:1 {theme_data.get('accent', theme_data['primary_light'])});
                    border: {border};
                    border-radius: 12px;
                    color: white;
                    font-size: 11px;
                    font-weight: 600;
                    padding: 6px;
                    text-align: center;
                }}
                QPushButton:hover {{
                    border: 3px solid {theme_data['primary_light']};
                }}
            """)
            btn.setText(("✓  " if is_active else "") + theme_data['display_name'])

        mw = self.window()
        if hasattr(mw, 'apply_theme'):
            mw.apply_theme(theme_key)

    def _save_calendar_mode(self):
        mode = self.cal_mode_cb.currentData()
        self.db.save_setting('calendar_mode', mode)
        # Propagate to calendar page via main window
        mw = self.window()
        if hasattr(mw, 'calendar_page'):
            mw.calendar_page.set_hebrew_mode(mode == 'hebrew')
        QMessageBox.information(self, "הצלחה", "הגדרת לוח השנה נשמרה!")

    def _backup_db(self):
        path, _ = QFileDialog.getSaveFileName(self, "שמור גיבוי", f"hashala_backup_{date.today()}.db", "DB (*.db)")
        if path:
            import shutil
            shutil.copy2(str(DB_PATH), path)
            QMessageBox.information(self, "הצלחה", f"גיבוי נשמר:\n{path}")

    def refresh(self):
        pass


# ─────────────────────────────────────────────
#  BACKGROUND REMINDER WORKER
# ─────────────────────────────────────────────

class ReminderWorker(QThread):
    reminder_needed = Signal(list)

    def __init__(self, db):
        super().__init__()
        self.db = db
        self._running = True

    def run(self):
        while self._running:
            try:
                self.db.update_overdue_statuses()
                settings = self.db.get_settings()
                days_ahead = int(settings.get('reminder_days', 2))
                target = (date.today() + timedelta(days=days_ahead)).isoformat()
                upcoming = self.db.conn.execute("""
                    SELECT l.*, i.name as item_name, b.full_name as borrower_name,
                           b.phone, b.email as borrower_email
                    FROM loans l JOIN items i ON l.item_id=i.id
                    JOIN borrowers b ON l.borrower_id=b.id
                    WHERE l.planned_return=? AND l.status='פעיל'
                """, (target,)).fetchall()
                if upcoming:
                    self.reminder_needed.emit([dict(r) for r in upcoming])
            except Exception:
                pass
            # Sleep 1 hour (check every hour)
            for _ in range(3600):
                if not self._running:
                    break
                self.msleep(1000)

    def stop(self):
        self._running = False


# ─────────────────────────────────────────────
#  MAIN WINDOW
# ─────────────────────────────────────────────

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = Database()
        self.setWindowTitle('השאלה חכמה - ניהול גמ"ח')
        self.setMinimumSize(1100, 700)
        self.resize(1500, 940)
        self.setLayoutDirection(Qt.RightToLeft)

        # Apply the saved theme (falls back to default if none saved yet)
        saved_theme = self.db.get_settings().get('theme', DEFAULT_THEME)
        self._current_theme = saved_theme if saved_theme in THEMES else DEFAULT_THEME
        global COLORS, STYLESHEET
        COLORS = dict(THEMES[self._current_theme])
        STYLESHEET = build_stylesheet(COLORS)
        self.setStyleSheet(STYLESHEET)

        self._setup_ui()
        self._setup_tray()
        self._start_worker()

    def apply_theme(self, theme_key):
        """Re-apply the global stylesheet and refresh HTML pages with a new theme."""
        global COLORS, STYLESHEET
        if theme_key not in THEMES:
            return
        self._current_theme = theme_key
        COLORS = dict(THEMES[theme_key])
        STYLESHEET = build_stylesheet(COLORS)
        self.setStyleSheet(STYLESHEET)
        # Re-render the HTML-based pages so their CSS variables pick up the new theme
        for page in [self.dashboard_page, self.reports_page, self.calendar_page]:
            if hasattr(page, 'refresh'):
                page.refresh()

    def _setup_ui(self):
        central = QWidget()
        central.setObjectName("centralWidget")
        self.setCentralWidget(central)

        main_layout = QHBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # ── Sidebar ──
        sidebar = QWidget()
        sidebar.setObjectName("sidebar")
        sidebar.setFixedWidth(220)
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(0, 0, 0, 0)
        sidebar_layout.setSpacing(0)

        # Logo / title
        logo_frame = QFrame()
        logo_frame.setObjectName("logoFrame")
        logo_layout = QVBoxLayout(logo_frame)
        logo_layout.setContentsMargins(16, 20, 16, 20)
        logo_lbl = QLabel('📦 השאלה חכמה')
        logo_lbl.setStyleSheet("color: white; font-size: 17px; font-weight: 700;")
        logo_layout.addWidget(logo_lbl)
        sub_lbl = QLabel('v1.5  |  ניהול גמ"ח חפצים')
        sub_lbl.setStyleSheet("color: #7FB3D3; font-size: 11px;")
        logo_layout.addWidget(sub_lbl)
        sidebar_layout.addWidget(logo_frame)

        # Nav buttons
        self.nav_buttons = []
        nav_items = [
            ("🏠  לוח בקרה", 0),
            ("📦  ניהול חפצים", 1),
            ("👥  ניהול שואלים", 2),
            ("📋  השאלות", 3),
            ("🔔  תזכורות", 4),
            ("📅  לוח שנה", 5),
            ("📊  דוחות", 6),
            ("⚙️  הגדרות", 7),
        ]
        for label, idx in nav_items:
            btn = QPushButton(label)
            btn.setObjectName("navBtn")
            btn.setMinimumHeight(48)
            btn.clicked.connect(lambda _, i=idx: self._navigate(i))
            sidebar_layout.addWidget(btn)
            self.nav_buttons.append(btn)

        sidebar_layout.addStretch()

        # Version
        ver = QLabel("v1.5")
        ver.setStyleSheet("color: #4A5568; font-size: 10px; padding: 8px 16px;")
        sidebar_layout.addWidget(ver)

        # ── Content ──
        self.pages = QStackedWidget()
        self.dashboard_page = DashboardPage(self.db)
        self.dashboard_page.navigate.connect(self._navigate)
        self.items_page = ItemsPage(self.db)
        self.borrowers_page = BorrowersPage(self.db)
        self.loans_page = LoansPage(self.db)
        self.new_loan_page = NewLoanPage(self.db)
        self.new_loan_page.loan_created.connect(self._on_loan_created)
        self.reminders_page = RemindersPage(self.db)
        self.calendar_page = CalendarPage(self.db)
        self.reports_page = ReportsPage(self.db)
        self.settings_page = SettingsPage(self.db)

        for p in [
            self.dashboard_page, self.items_page, self.borrowers_page,
            self.loans_page, self.reminders_page, self.calendar_page,
            self.reports_page, self.settings_page
        ]:
            self.pages.addWidget(p)

        # new_loan_page is added separately and navigated to via index 8
        self.pages.addWidget(self.new_loan_page)  # index 8

        main_layout.addWidget(sidebar)
        main_layout.addWidget(self.pages)

        self._navigate(0)

    def _navigate(self, idx):
        self.pages.setCurrentIndex(idx)
        # nav buttons cover indices 0-7; idx 8 = new_loan_page (no nav button)
        for i, btn in enumerate(self.nav_buttons):
            btn.setProperty("active", i == idx)
            btn.style().unpolish(btn)
            btn.style().polish(btn)
        page = self.pages.currentWidget()
        if hasattr(page, 'refresh'):
            page.refresh()

    def _on_loan_created(self):
        self.dashboard_page.refresh()
        self.loans_page.refresh()
        self.reminders_page.refresh()

    def _setup_tray(self):
        try:
            self.tray = QSystemTrayIcon(self)
            pix = QPixmap(32, 32)
            pix.fill(QColor(COLORS['primary_light']))
            self.tray.setIcon(QIcon(pix))
            tray_menu = QMenu()
            show_action = QAction("פתח את האפליקציה", self)
            show_action.triggered.connect(self.show)
            quit_action = QAction("יציאה", self)
            quit_action.triggered.connect(QApplication.quit)
            tray_menu.addAction(show_action)
            tray_menu.addSeparator()
            tray_menu.addAction(quit_action)
            self.tray.setContextMenu(tray_menu)
            self.tray.setToolTip('השאלה חכמה - גמ"ח')
            self.tray.show()
        except Exception:
            pass

    def _start_worker(self):
        self.worker = ReminderWorker(self.db)
        self.worker.reminder_needed.connect(self._show_reminders)
        self.worker.start()

    def _show_reminders(self, loans):
        names = ", ".join(l['item_name'] for l in loans[:3])
        try:
            self.tray.showMessage(
                "תזכורת השאלה חכמה",
                f"יש {len(loans)} חפצים שצריך להחזיר בקרוב: {names}",
                QSystemTrayIcon.Information,
                5000
            )
        except Exception:
            pass

    def closeEvent(self, event):
        if hasattr(self, 'tray') and self.tray.isVisible():
            self.hide()
            event.ignore()
        else:
            if hasattr(self, 'worker'):
                self.worker.stop()
            event.accept()


# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────

def main():
    app = QApplication(sys.argv)
    app.setApplicationName('השאלה חכמה')
    app.setApplicationVersion("1.5")
    app.setLayoutDirection(Qt.RightToLeft)

    # Apply global font
    font = QFont("Segoe UI", 10)
    app.setFont(font)

    window = MainWindow()
    # Open at a generous, screen-relative default size so the dashboard's
    # HTML/CSS layout and all data tables have room to breathe.
    screen = app.primaryScreen()
    if screen:
        avail = screen.availableGeometry()
        target_w = min(int(avail.width() * 0.88), 1680)
        target_h = min(int(avail.height() * 0.88), 1000)
        window.resize(max(target_w, 1200), max(target_h, 760))
        # Center on screen
        x = avail.x() + (avail.width() - window.width()) // 2
        y = avail.y() + (avail.height() - window.height()) // 2
        window.move(x, y)
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
